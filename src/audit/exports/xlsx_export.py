"""Excel (.xlsx) audit report, the whole report, every section a table.

A multi-sheet workbook that turns the Markdown audit report into something
you can sort, filter, and track in a spreadsheet:

* **Summary**, the at-a-glance dashboard: scan metadata plus every headline
  rollup (counts, severity, WCAG level, POUR principle, coverage).
* **Issues Overview**, the ticket index: a metadata header block plus one
  sortable row per issue (ID · Issue · Severity · Conformance Level ·
  Remediation Ownership · Status · Instances · Pages), each linking into that
  issue's own tab.
* **One tab per issue**, the ticket itself: a Current Behavior / Severity /
  Impact / Should be / Instances / References summary block, then one row per
  place the defect was actually seen (Where · User action · Element · What to
  fix · How to reproduce), the acceptance criteria, and a circled location
  screenshot when a ``blob_store`` is supplied. Past ``_MAX_ISSUE_SHEETS``
  issues the index still lists everything and the remaining instances pool
  into a single **More Issues** sheet.

  Fields a scan cannot establish, the component framework and assistive tech
  used, a fix-approach category, competing fix options with trade-offs, are
  omitted rather than emitted blank or inferred. Everything in the workbook
  came out of the scan.
* **Page Hotspots**, pages ranked by a severity-weighted load.
* **Page References**, every affected page as a clickable link, with a
  plain-language element location and the retained technical target.
* **Who's Affected**, open issues by the user ability each one blocks.
* **Coverage & Method**, per-criterion automated-vs-manual coverage.
* **Test Tracking**, the manual Pass / Fail checklist, one row per WCAG 2.2
  A/AA criterion.
* **Manual Review Evidence**, human outcomes, rationale, page references,
  and external evidence links kept with the workbook handoff.

Every issue-derived sheet is built from :func:`audit_report.build_audit_cards`
— the exact card model the Markdown audit report renders, so the two
deliverables can't drift. The coverage/checklist sheets come from
``coverage_matrix``. All of it is pinned by tests.
"""

from __future__ import annotations

import io
import json
import re
import sqlite3
from dataclasses import dataclass
from datetime import datetime
from typing import Any

from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.hyperlink import Hyperlink
from openpyxl.worksheet.worksheet import Worksheet

from audit import coverage_matrix, evaluation
from audit.blob_store import BlobStore
from audit.exports import interaction_coverage
from audit.exports.audit_report import (
    _IMPACT_SEVERITY,
    _PIPELINE_LABEL,
    MAX_HOTSPOTS,
    AuditCard,
    DroppedFinding,
    FixOption,
    IssueLocation,
    _principle_for,
    build_audit_cards,
    fix_options_for,
    issue_locations,
    load_report_rules,
)
from audit.exports.collector import ExportScan
from audit.logging import get_logger
from audit.web import issues

log = get_logger(__name__)

# Embedded-evidence sizing: scale screenshots down to this pixel width so a
# wide element capture doesn't blow out the Evidence column; the height is
# scaled proportionally and the row grown to fit.
_EVIDENCE_MAX_WIDTH = 240

# --- palette (matches the source template) --------------------------------
_HEADER_FILL = PatternFill("solid", fgColor="FF356854")  # dark teal-green
_HEADER_FONT = Font(bold=True, color="FFFFFFFF", size=11)
_META_FILL = PatternFill("solid", fgColor="FFFFF2CC")  # light yellow
_META_LABEL_FONT = Font(bold=True, size=11)
_BAND_FILL = PatternFill("solid", fgColor="FFF6F8F9")  # zebra stripe
_TITLE_FONT = Font(bold=True, size=14, color="FF356854")
_WRAP_TOP = Alignment(wrap_text=True, vertical="top")
_WRAP_TOP_CENTER = Alignment(wrap_text=True, vertical="top", horizontal="center")
_THIN = Side(style="thin", color="FFD0D7DE")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

# --- value mappings -------------------------------------------------------
# The template grades conformance A / AA / S, where "S" = a best-practice
# recommendation beyond the WCAG success criteria. Axcess labels those "BP".
_CONFORMANCE_DISPLAY = {"A": "A", "AA": "AA", "AAA": "AAA", "BP": "S"}
_STATUS_DISPLAY = {
    "new": "Not Started",
    "reviewing": "Reviewing",
    "in_progress": "In Progress",
    "remediated": "Resolved",
    "accepted_risk": "Accepted Risk",
    "false_positive": "Not an Issue",
}
_STATUS_CHOICES = [
    "Not Started",
    "Reviewing",
    "In Progress",
    "Resolved",
    "Accepted Risk",
    "Not an Issue",
]
_OWNER_DISPLAY = {
    "dev": "Developer",
    "editor": "Editor",
    "designer": "Designer",
    "content": "Content author",
}

_GUIDANCE = (
    "Prioritization Guidance, Start with rows whose Evidence decision is "
    "Likely Barrier. Rows marked Expert Review require a human decision before "
    "they are described as barriers; Informational rows are not worklist items. "
    "Within the Likely Barrier lane, fix issues in conformance-level order. "
    "Level A flags the most significant barriers that block people with "
    "disabilities outright; remediate these first. Level AA issues are "
    "required for WCAG 2.2 AA conformance (the standard most policies and "
    "laws reference) and should follow. Items marked S are best-practice "
    "recommendations that go beyond the success criteria, address them "
    "once the A and AA issues are resolved. Within a level, work top-down: "
    "the table is sorted by Axcess's priority score (occurrence count, "
    "impact, and reach)."
)

# The index: one row per issue, sortable, each linking to that issue's tab.
_ISSUE_HEADERS = (
    "ID",
    "Issue",
    "Severity",
    "Conformance Level",
    "Remediation Ownership",
    "Status",
    "Instances",
    "Pages",
    "Details",
)
_ISSUE_WIDTHS = (8.0, 46.0, 12.0, 17.5, 19.13, 15.38, 11.0, 9.0, 22.0)

# One issue's own tab: the instance table that the index links into.
_INSTANCE_HEADERS = (
    "#",
    "Where",
    "User action",
    "Element",
    "What to fix",
    "How to reproduce",
)
_INSTANCE_WIDTHS = (24.0, 34.0, 26.0, 40.0, 46.0, 34.0)

# Excel hard-limits sheet names to 31 characters and forbids these.
_SHEET_NAME_BANNED = str.maketrans(dict.fromkeys("[]:*?/\\", "-"))
# One tab per issue is the readable layout, but a workbook with a tab per
# issue on a large scan is not openable in practice. Past this many, the
# index still lists every issue and the remainder's instances are pooled
# into one sheet rather than silently dropped.
_MAX_ISSUE_SHEETS = 40
_POOLED_SHEET = "More Issues"
_POOLED_HEADERS = ("ID", "Issue", "#", "Where", "User action", "Element", "What to fix")
_POOLED_WIDTHS = (8.0, 34.0, 5.0, 30.0, 24.0, 36.0, 42.0)

_TRACK_HEADERS = ("Focus", "What to check", "Pass / Fail")
_TRACK_WIDTHS = (28.0, 70.0, 14.0)
_PASS_FAIL_CHOICES = ["Pass", "Fail", "Needs follow-up", "N/A", "Not tested"]

_URL_IN_TEXT = re.compile(r"https?://[^\s]+")

_LOCATION_CAP = 6  # page links listed before collapsing to "+N more"

# Most-urgent-first ordering for the severity rollups.
_SEVERITY_ORDER = ("Critical", "Serious", "Moderate", "Minor")
# Page-hotspot weighting, mirrors audit_report._page_hotspots so the Excel
# "Page Hotspots" sheet ranks pages identically to the Markdown report.
_SEVERITY_WEIGHT = {"Critical": 4.0, "Serious": 3.0, "Moderate": 2.0, "Minor": 1.0}
# User-group labels, mirrors audit_report._abilities_rollup.
_ABILITY_LABEL = {
    "vision": "Vision (blind / low-vision / color-blind)",
    "motor": "Motor (keyboard-only / switch / tremor)",
    "cognition": "Cognition (memory / attention / language)",
    "hearing": "Hearing (deaf / hard-of-hearing)",
}


def _fmt_date(raw: str | None) -> str:
    """ISO timestamp → 'June 24, 2026'; falls back to the raw string."""
    if not raw:
        return ""
    try:
        return datetime.fromisoformat(raw.replace("Z", "+00:00")).strftime("%B %-d, %Y")
    except (ValueError, TypeError):
        return raw


def _dominant_status(summary: dict[str, int]) -> str:
    """The status bucket with the most occurrences → its display label."""
    if not summary:
        return _STATUS_DISPLAY["new"]
    bucket = max(summary, key=lambda k: summary.get(k, 0))
    return _STATUS_DISPLAY.get(bucket, "Not Started")


def _bullets(steps: tuple[str, ...] | list[str]) -> str:
    """Render fix steps as a newline-bulleted block for one cell."""
    items = [s.strip() for s in steps if s and s.strip()]
    return "\n".join(f"• {s}" for s in items)


def _locations(pages: list[issues.IssuePage], total_pages: int) -> str:
    """A concise link to the separate sheet of individually clickable pages."""
    urls = [p.page_url for p in pages if p.page_url]
    if not urls:
        return "No page-level location was recorded."
    # Site-wide problem (every crawled page) reads better as a phrase.
    if total_pages > 1 and len(set(urls)) >= total_pages:
        return "All scanned pages, open Page References."
    return f"{len(set(urls))} linked page(s), open Page References."


def _link_first_url(cell: Any) -> None:
    """Make a URL inside a cell's text clickable.

    Longform cells (fix steps, reproduction notes) can carry a rule-docs link
    inline. A recipient must never have to select and copy a visible URL out
    of a spreadsheet, so the cell itself carries the link.
    """
    if cell.hyperlink is not None:
        return
    match = _URL_IN_TEXT.search(str(cell.value or ""))
    if match:
        cell.hyperlink = match.group(0).rstrip(").,;")


def _style_header_row(ws: Worksheet, row: int, ncols: int) -> None:
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = _WRAP_TOP
        cell.border = _BORDER


def _embed_evidence(
    ws: Worksheet,
    *,
    row: int,
    col: int,
    detail: issues.IssueDetail | None,
    blob_store: BlobStore,
) -> None:
    """Embed the first available finding screenshot into the row's Evidence cell.

    Pulls the first screenshot hash across ``detail.pages``, resolves it to a
    PNG in the blob store, scales it to fit the column, and anchors it to the
    Evidence cell, growing the row to fit. A missing or corrupt file never
    breaks the export (the whole body is defensive); the cell stays blank.
    """
    if detail is None:
        return
    hash_ = next(
        (h for p in detail.pages for h in p.screenshot_hashes if h),
        None,
    )
    if not hash_:
        return
    try:
        path = blob_store.path_for(BlobStore._rel_path(hash_, "png"))
        if not path.exists():
            return
        img = XLImage(str(path))
        if img.width and img.width > _EVIDENCE_MAX_WIDTH:
            img.height = int(img.height * _EVIDENCE_MAX_WIDTH / img.width)
            img.width = _EVIDENCE_MAX_WIDTH
        col_letter = get_column_letter(col)
        img.anchor = f"{col_letter}{row}"
        ws.add_image(img)
        existing = ws.row_dimensions[row].height or 0
        ws.row_dimensions[row].height = max(existing, img.height * 0.75)
    except Exception as exc:  # pragma: no cover - defensive
        log.warning("xlsx.evidence_embed_failed", hash=hash_, error=str(exc))


@dataclass(frozen=True)
class _Ticket:
    """One issue as the workbook presents it: a summary plus its instances.

    Built from :class:`issues.IssueRow`, not from ``AuditCard``: the card
    model deliberately buckets best-practice and already-triaged issues out
    of the main list, and an index that quietly omitted them would not be an
    index of the scan.
    """

    index: int
    row: issues.IssueRow
    detail: issues.IssueDetail | None
    locations: list[IssueLocation]
    overflow: int
    #: Authored fix approaches for this rule. Empty for rules nobody has
    #: written options for, and the sheet then omits the section rather
    #: than printing a table with one blank row.
    fix_options: tuple[FixOption, ...] = ()

    @property
    def id(self) -> str:
        return f"I{self.index + 1:02d}"

    @property
    def category(self) -> str:
        """The WCAG principle this issue's criterion belongs to.

        Reuses the Summary sheet's grouping rather than inventing a second
        taxonomy, so a reader who filters the overview by category gets the
        same buckets the dashboard counted.
        """
        return _principle_for(self.row.wcag_sc)

    @property
    def severity(self) -> str:
        return _IMPACT_SEVERITY.get(self.row.impact or "", "Moderate")

    @property
    def current_behavior(self) -> str:
        detail = (self.detail.description if self.detail else None) or self.row.description
        return detail or self.row.title

    @property
    def impact(self) -> str:
        why = (self.detail.why_matters if self.detail else None) or self.row.why_matters
        return why or "Users relying on assistive technology hit a barrier here."

    @property
    def should_be(self) -> str | None:
        return (self.detail.acceptance if self.detail else None) or self.row.acceptance

    @property
    def fix_steps(self) -> list[str]:
        steps = (self.detail.fix_steps if self.detail else None) or list(self.row.fix_steps)
        return list(steps)

    @property
    def reproduce(self) -> str:
        if self.detail is None:
            return ""
        return self.detail.verify_manual or self.detail.verify_automated or ""

    @property
    def help_url(self) -> str:
        return (self.detail.help_url if self.detail else None) or self.row.help_url or ""

    @property
    def references(self) -> str:
        parts = [
            f"{self.row.wcag_sc} {self.row.wcag_name}".strip() if self.row.wcag_sc else "",
            f"Level {self.row.conformance}" if self.row.conformance in ("A", "AA", "AAA") else "",
        ]
        return " · ".join(part for part in parts if part)


_ENVIRONMENT_TO_SUPPLY = (
    "Component framework and assistive technology tested:, (add before sharing)"
)


def _scan_environment(scan: ExportScan, conn: sqlite3.Connection) -> str:
    """Describe the conditions the scan actually ran under.

    Only what the scan can prove: which engines produced findings, whether
    pages were browser-rendered, the viewport, and when. The component
    framework and the assistive tech a human tested with are not knowable
    from a crawl, so the line ends with an explicit slot for them rather
    than being left off (a reader would not know it was missing) or guessed
    (a reader would believe it).
    """
    engines = []
    if scan.axe_pages_scanned:
        engines.append(f"axe-core ({scan.axe_level})")
    if scan.alfa_pages_scanned:
        engines.append("Alfa")
    config: dict[str, Any] = {}
    try:
        raw = conn.execute("SELECT config_json FROM scans WHERE id = ?", (scan.id,)).fetchone()
        loaded = json.loads((raw[0] if raw else None) or "{}")
        config = loaded if isinstance(loaded, dict) else {}
    except (sqlite3.Error, json.JSONDecodeError, TypeError, IndexError):
        config = {}
    rendering = "Browser-rendered" if config.get("js_eager") else "Static HTML (no browser)"
    viewport = config.get("viewport") or config.get("browser_viewport")
    parts = [
        f"Scan #{scan.id} of {scan.seed_url}",
        f"{scan.page_count} page(s) crawled",
        rendering,
    ]
    if isinstance(viewport, str | list | tuple) and viewport:
        parts.append(f"Viewport {viewport}")
    if engines:
        parts.append("Engines: " + ", ".join(engines))
    finished = _fmt_date(scan.finished_at or scan.started_at)
    if finished:
        parts.append(f"Scanned {finished}")
    return ". ".join(parts) + ".\n" + _ENVIRONMENT_TO_SUPPLY


def _build_tickets(conn: sqlite3.Connection, scan: ExportScan) -> list[_Ticket]:
    # Read the authored rule copy once for the whole workbook, not once per
    # issue: a scan with 200 issue types would otherwise re-parse the YAML
    # 200 times to answer the same question.
    rules = load_report_rules()
    tickets: list[_Ticket] = []
    for index, row in enumerate(issues.list_issues(conn, scan.id)):
        locations, overflow = issue_locations(conn, row)
        tickets.append(
            _Ticket(
                index=index,
                row=row,
                detail=issues.get_issue_detail(conn, scan.id, row.issue_key),
                locations=locations,
                overflow=overflow,
                fix_options=fix_options_for(row, rules),
            )
        )
    return tickets


def _sheet_name(ticket: _Ticket, taken: set[str]) -> str:
    """A unique, Excel-legal tab name that still reads as the issue.

    Excel caps names at 31 characters and rejects ``[]:*?/\\``. The ID prefix
    survives whatever gets truncated, because that is what the index column
    shows and what the reader matches against.
    """
    prefix = ticket.id
    room = 31 - len(prefix) - 1
    cleaned = " ".join(str(ticket.row.title or "Issue").translate(_SHEET_NAME_BANNED).split())
    name = f"{prefix} {cleaned[:room]}".strip()
    if name not in taken:
        taken.add(name)
        return name
    # Two titles can collide once truncated; the ID alone cannot.
    for suffix in range(2, 100):
        candidate = f"{prefix} {cleaned[: max(0, room - 3)]}~{suffix}".strip()
        if candidate not in taken:
            taken.add(candidate)
            return candidate
    taken.add(prefix)
    return prefix


def _instance_rows(ticket: _Ticket) -> list[tuple[Any, ...]]:
    """One row per place the defect was actually seen.

    ``User action`` is only meaningful for findings the interaction probe
    revealed: those record the control that had to be operated to expose the
    markup. Everything else was present at load, and says so rather than
    inventing a journey.
    """
    fix = _bullets(ticket.fix_steps)
    reproduce = ticket.reproduce
    rows: list[tuple[Any, ...]] = []
    for number, loc in enumerate(ticket.locations, start=1):
        where = loc.page_title or loc.page_url or "n/a"
        action = f'Open "{loc.revealed_by}" on this page.' if loc.revealed_by else "Load the page."
        element = loc.selector or loc.image_url or loc.description or "n/a"
        rows.append((number, where, action, element, fix, reproduce))
    return rows


def _build_issue_index_sheet(
    ws: Worksheet,
    scan: ExportScan,
    tickets: list[_Ticket],
    sheet_names: list[str | None],
    *,
    auditor: str,
    audit_date: str,
) -> None:
    """The ticket index: every issue, each linking into its own tab."""
    ncols = len(_ISSUE_HEADERS)
    for i, w in enumerate(_ISSUE_WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    meta = [
        ("Page", scan.seed_url or "n/a"),
        ("Compliance Standard", f"WCAG 2.2 Level {scan.axe_level}"),
        ("Audit Date", audit_date),
        ("Auditor", auditor),
        ("Prioritization Guidance", _GUIDANCE),
    ]
    for r, (label, value) in enumerate(meta, start=1):
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=ncols)
        cell = ws.cell(row=r, column=1, value=f"{label}: {value}")
        cell.fill = _META_FILL
        cell.alignment = _WRAP_TOP
        cell.font = _META_LABEL_FONT if label != "Prioritization Guidance" else Font(size=10)
        if value.startswith(("https://", "http://")):
            cell.hyperlink = value
            cell.style = "Hyperlink"
    ws.row_dimensions[5].height = 96

    header_row = 8
    for c, name in enumerate(_ISSUE_HEADERS, start=1):
        ws.cell(row=header_row, column=c, value=name)
    _style_header_row(ws, header_row, ncols)

    r = header_row
    for ticket in tickets:
        r += 1
        target = sheet_names[ticket.index]
        values: tuple[Any, ...] = (
            ticket.id,
            ticket.row.title,
            ticket.severity,
            _CONFORMANCE_DISPLAY.get(ticket.row.conformance, ticket.row.conformance),
            _OWNER_DISPLAY.get(ticket.row.responsibility, ticket.row.responsibility.title()),
            _dominant_status(ticket.row.status_summary),
            ticket.row.occurrence_count,
            ticket.row.page_count,
            "Open instances" if target else f"See {_POOLED_SHEET}",
        )
        for c, v in enumerate(values, start=1):
            cell = ws.cell(row=r, column=c, value=v)
            cell.alignment = _WRAP_TOP_CENTER if c in (1, 3, 4, 5, 6, 7, 8) else _WRAP_TOP
            cell.border = _BORDER
            if ticket.index % 2 == 1:
                cell.fill = _BAND_FILL
        link_cell = ws.cell(row=r, column=ncols)
        link_cell.hyperlink = Hyperlink(
            ref=link_cell.coordinate,
            location=f"'{target or _POOLED_SHEET}'!A1",
            display=str(link_cell.value or ""),
            tooltip="Open this issue's instances",
        )
        link_cell.style = "Hyperlink"

    last_row = max(r, header_row + 1)
    ws.auto_filter.ref = f"A{header_row}:{get_column_letter(ncols)}{last_row}"
    ws.freeze_panes = f"A{header_row + 1}"

    dv = DataValidation(type="list", formula1=f'"{",".join(_STATUS_CHOICES)}"', allow_blank=True)
    ws.add_data_validation(dv)
    if r > header_row:
        dv.add(f"F{header_row + 1}:F{last_row}")


_FIX_OPTION_HEADERS = ("Option", "Applies to", "Approach", "Watch out for")


def _write_fix_options(ws: Worksheet, ticket: _Ticket, *, row: int, ncols: int) -> int:
    """Render the authored fix approaches, or nothing at all.

    Returns the last row written. The section is skipped when a rule has no
    authored options: an empty Option/Approach table would read as "there is
    no way to fix this", and a table with one row silently invented from the
    fix steps would read as "these were the alternatives considered". Both
    are claims the scan cannot support.

    "Applies to" names the scope rather than repeating a per-instance list:
    options are authored per rule, so they apply to every instance of it,
    and a column that said "Instance 1" on an issue with 300 of them would
    be wrong in 299 places.
    """
    if not ticket.fix_options:
        return row - 2

    heading = ws.cell(row=row, column=1, value="Suggested fix")
    heading.font = _META_LABEL_FONT
    row += 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    note = ws.cell(
        row=row,
        column=1,
        value=(
            "Approaches shared for context, not prescribed. Use what fits the "
            "codebase, as long as the acceptance criteria hold."
        ),
    )
    note.font = Font(italic=True, size=10)
    note.alignment = _WRAP_TOP

    row += 1
    header_row = row
    for c, name in enumerate(_FIX_OPTION_HEADERS, start=1):
        ws.cell(row=header_row, column=c, value=name)
    _style_header_row(ws, header_row, len(_FIX_OPTION_HEADERS))

    count = ticket.row.occurrence_count
    applies = "All instances" if count == 1 else f"All {count} instances"
    for idx, option in enumerate(ticket.fix_options):
        row += 1
        values = (
            f"{chr(ord('A') + idx)}. {option.label}",
            applies,
            option.approach,
            option.watch_out or "n/a",
        )
        for c, value in enumerate(values, start=1):
            cell = ws.cell(row=row, column=c, value=value)
            cell.alignment = _WRAP_TOP
            cell.border = _BORDER
            if idx % 2 == 1:
                cell.fill = _BAND_FILL
    return row


def _build_issue_detail_sheet(
    ws: Worksheet,
    ticket: _Ticket,
    *,
    environment: str,
    blob_store: BlobStore | None,
) -> None:
    """One issue's tab: what it is, how to fix it, then every instance.

    Three fields here are not scan output, and each is handled so a reader
    can tell which is which. ``Category`` is derived from the criterion
    rather than authored. ``Environment`` states the conditions the crawl
    can prove and leaves a marked slot for the framework and assistive tech
    only a human knows. ``Suggested fix`` renders approaches written into
    ``rules/audit_report.yaml``, and the section is omitted entirely for
    rules that have none. Nothing is inferred to fill out a table.
    """
    for i, w in enumerate(_INSTANCE_WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ncols = len(_INSTANCE_HEADERS)
    last_col = get_column_letter(ncols)

    ws.merge_cells(f"A1:{last_col}1")
    title = ws.cell(row=1, column=1, value=f"{ticket.id} · {ticket.row.title}")
    title.font = _TITLE_FONT

    back = ws.cell(row=2, column=1, value="← Back to Issues Overview")
    back.hyperlink = Hyperlink(
        ref=back.coordinate,
        location="'Issues Overview'!A1",
        display="← Back to Issues Overview",
    )
    back.style = "Hyperlink"

    summary: list[tuple[str, Any]] = [
        ("Current Behavior", ticket.current_behavior),
        ("Severity", ticket.severity),
        ("Category", ticket.category),
        ("Impact", ticket.impact),
    ]
    if ticket.should_be:
        # One field, one place. Axcess stores a single acceptance statement
        # per rule, so printing it as both "Should be" and a separate
        # "Acceptance criteria" section put the same sentence on the sheet
        # twice and implied two sources that do not exist.
        summary.append(("Should be (acceptance criteria)", ticket.should_be))
    summary.append(("Instances", ticket.row.occurrence_count))
    summary.append(("Pages", ticket.row.page_count))
    summary.append(("Environment", environment))
    if ticket.references:
        summary.append(("References", ticket.references))
    if ticket.help_url:
        summary.append(("Rule documentation", ticket.help_url))

    r = 3
    for label, value in summary:
        r += 1
        key = ws.cell(row=r, column=1, value=label)
        key.font = _META_LABEL_FONT
        key.alignment = _WRAP_TOP
        key.fill = _META_FILL
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=ncols)
        cell = ws.cell(row=r, column=2, value=value)
        cell.alignment = _WRAP_TOP
        _link_first_url(cell)

    r += 2
    # Not just "Instances": that word is already a field in the block above,
    # and a reader scanning column A should not meet it twice meaning two
    # different things.
    heading = ws.cell(row=r, column=1, value="Instances, one row per occurrence found")
    heading.font = _META_LABEL_FONT
    r += 1
    header_row = r
    for c, name in enumerate(_INSTANCE_HEADERS, start=1):
        ws.cell(row=header_row, column=c, value=name)
    _style_header_row(ws, header_row, ncols)

    rows = _instance_rows(ticket)
    for idx, values in enumerate(rows):
        r += 1
        for c, v in enumerate(values, start=1):
            cell = ws.cell(row=r, column=c, value=v)
            cell.alignment = _WRAP_TOP_CENTER if c == 1 else _WRAP_TOP
            cell.border = _BORDER
            _link_first_url(cell)
            if idx % 2 == 1:
                cell.fill = _BAND_FILL
    if not rows:
        r += 1
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=ncols)
        note = ws.cell(row=r, column=1, value="No retained instance locations for this issue.")
        note.font = Font(italic=True, size=10)
    if ticket.overflow:
        r += 1
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=ncols)
        # Say what is not listed. A table of ten under a heading claiming 300
        # occurrences is the quiet gap this workbook exists to avoid.
        overflow = ws.cell(
            row=r,
            column=1,
            value=(
                f"+{ticket.overflow} further occurrence(s) recorded for this issue are "
                "not listed individually here."
            ),
        )
        overflow.font = Font(italic=True, size=10)
        overflow.alignment = _WRAP_TOP

    r = _write_fix_options(ws, ticket, row=r + 2, ncols=ncols)

    if blob_store is not None:
        _embed_evidence(
            ws, row=header_row - 1, col=ncols, detail=ticket.detail, blob_store=blob_store
        )


def _build_pooled_instances_sheet(ws: Worksheet, pooled: list[_Ticket]) -> None:
    """Instances for issues past the per-issue tab cap, in one flat table."""
    rows: list[tuple[Any, ...]] = []
    for ticket in pooled:
        for values in _instance_rows(ticket):
            rows.append((ticket.id, ticket.row.title, *values[:4], values[4]))
    _styled_table(
        ws,
        title="More Issues",
        subtitle=(
            "Instances for issues beyond the per-issue tab limit. Every issue is "
            "still listed on Issues Overview; only the tab is pooled here."
        ),
        headers=_POOLED_HEADERS,
        rows=rows,
        widths=_POOLED_WIDTHS,
        center_cols=(1, 3),
        empty_note="No pooled issues for this scan.",
    )


def _build_tracking_sheet(ws: Worksheet, *, manual_outcomes: dict[str, str] | None = None) -> None:
    for i, w in enumerate(_TRACK_WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.merge_cells("A1:C1")
    title = ws.cell(row=1, column=1, value="Manual Test Tracking")
    title.font = _TITLE_FONT
    ws.merge_cells("A2:C2")
    sub = ws.cell(
        row=2,
        column=1,
        value=(
            "Automated checks can't decide every criterion. For each WCAG 2.2 "
            "A/AA success criterion below, confirm by hand what the tool flags "
            "as needing a human, then record Pass / Fail."
        ),
    )
    sub.alignment = _WRAP_TOP
    sub.font = Font(size=10)
    ws.row_dimensions[2].height = 42

    header_row = 4
    for c, name in enumerate(_TRACK_HEADERS, start=1):
        ws.cell(row=header_row, column=c, value=name)
    _style_header_row(ws, header_row, len(_TRACK_HEADERS))

    crit = sorted(
        coverage_matrix.load_matrix(),
        key=lambda c: [int(p) for p in c.sc.split(".")],
    )
    r = header_row
    for idx, cr in enumerate(crit):
        r += 1
        values = (
            f"{cr.sc} {cr.name} (Level {cr.level})",
            cr.manual_check,
            (manual_outcomes or {}).get(cr.sc, ""),
        )
        for col, v in enumerate(values, start=1):
            cell = ws.cell(row=r, column=col, value=v)
            cell.alignment = _WRAP_TOP_CENTER if col == 3 else _WRAP_TOP
            cell.border = _BORDER
            if idx % 2 == 1:
                cell.fill = _BAND_FILL

    last_row = max(r, header_row + 1)
    ws.auto_filter.ref = f"A{header_row}:C{last_row}"
    ws.freeze_panes = f"A{header_row + 1}"

    dv = DataValidation(type="list", formula1=f'"{",".join(_PASS_FAIL_CHOICES)}"', allow_blank=True)
    ws.add_data_validation(dv)
    if r > header_row:
        dv.add(f"C{header_row + 1}:C{last_row}")


# --- generic data-sheet layout (used by every rollup sheet) ---------------


def _styled_table(
    ws: Worksheet,
    *,
    title: str,
    subtitle: str,
    headers: tuple[str, ...],
    rows: list[tuple[Any, ...]],
    widths: tuple[float, ...],
    center_cols: tuple[int, ...] = (),
    empty_note: str = "Nothing to report for this scan.",
) -> None:
    """Lay out one self-contained data sheet.

    Title (row 1) · subtitle (row 2) · a dark-green header (row 4) · banded,
    bordered data rows from row 5, with a frozen header and an auto-filter.
    Keeps every rollup sheet visually consistent.
    """
    ncols = len(headers)
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    last_col = get_column_letter(ncols)

    ws.merge_cells(f"A1:{last_col}1")
    tcell = ws.cell(row=1, column=1, value=title)
    tcell.font = _TITLE_FONT
    ws.merge_cells(f"A2:{last_col}2")
    scell = ws.cell(row=2, column=1, value=subtitle)
    scell.alignment = _WRAP_TOP
    scell.font = Font(size=10)
    ws.row_dimensions[2].height = 42

    header_row = 4
    for c, name in enumerate(headers, start=1):
        ws.cell(row=header_row, column=c, value=name)
    _style_header_row(ws, header_row, ncols)

    center = set(center_cols)
    r = header_row
    if not rows:
        ws.merge_cells(
            start_row=header_row + 1, start_column=1, end_row=header_row + 1, end_column=ncols
        )
        note = ws.cell(row=header_row + 1, column=1, value=empty_note)
        note.alignment = _WRAP_TOP
        note.font = Font(italic=True, size=10)
    else:
        for idx, row in enumerate(rows):
            r += 1
            for col, v in enumerate(row, start=1):
                cell = ws.cell(row=r, column=col, value=v)
                cell.alignment = _WRAP_TOP_CENTER if col in center else _WRAP_TOP
                cell.border = _BORDER
                _link_first_url(cell)
                if idx % 2 == 1:
                    cell.fill = _BAND_FILL
    last_row = max(r, header_row + 1)
    ws.auto_filter.ref = f"A{header_row}:{last_col}{last_row}"
    ws.freeze_panes = f"A{header_row + 1}"


# --- summary dashboard ----------------------------------------------------


def _build_summary_sheet(
    ws: Worksheet,
    scan: ExportScan,
    cards: list[AuditCard],
    dropped: list[DroppedFinding],
    best_practice: list[Any],
    *,
    audit_date: str,
    evaluation_record: dict[str, Any],
) -> None:
    """A one-screen dashboard: scan metadata + every headline rollup."""
    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 58

    ws.merge_cells("A1:B1")
    title = ws.cell(row=1, column=1, value=f"Accessibility audit, Scan #{scan.id}")
    title.font = _TITLE_FONT
    ws.merge_cells("A2:B2")
    sub = ws.cell(
        row=2,
        column=1,
        value=(
            "At-a-glance summary of this scan. Each rollup below has its own "
            "sheet with the detail; this is the dashboard for tracking progress."
        ),
    )
    sub.alignment = _WRAP_TOP
    sub.font = Font(size=10)
    ws.row_dimensions[2].height = 36

    r = 4

    def section(label: str) -> None:
        nonlocal r
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
        cell = ws.cell(row=r, column=1, value=label)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = _WRAP_TOP
        r += 1

    def metric(label: str, value: Any) -> Any:
        nonlocal r
        lc = ws.cell(row=r, column=1, value=label)
        lc.font = Font(bold=True, size=11)
        lc.alignment = _WRAP_TOP
        lc.border = _BORDER
        vc = ws.cell(row=r, column=2, value=value)
        vc.alignment = _WRAP_TOP
        vc.border = _BORDER
        r += 1
        return vc

    section("Scan")
    seed_url_cell = metric("Seed URL", scan.seed_url or "n/a")
    if scan.seed_url.startswith(("https://", "http://")):
        seed_url_cell.hyperlink = scan.seed_url
        seed_url_cell.style = "Hyperlink"
    metric("Audited against", f"WCAG 2.2 Level {scan.axe_level}")
    metric("Pages crawled", scan.page_count)
    metric("Audit date", audit_date)
    r += 1

    if evaluation_record["exists"]:
        section("Expert evaluation")
        metric(
            "Target",
            f"{evaluation_record['target_standard']} Level {evaluation_record['target_level']}",
        )
        metric("Review status", str(evaluation_record["status"]).replace("_", " "))
        for label, key in (
            ("Reviewer", "reviewer"),
            ("Purpose", "purpose"),
            ("Included scope", "scope_included"),
            ("Excluded scope", "scope_excluded"),
            ("Sample", "sample_description"),
            ("Methods", "methods_note"),
            ("Limitations", "limitations"),
        ):
            if evaluation_record[key]:
                metric(label, evaluation_record[key])
        r += 1

    likely_occurrences = sum(c.affected_finding_count for c in cards)
    review_occurrences = sum(int(row.occurrence_count) for row in best_practice)
    section("Issue counts")
    metric("Likely-barrier issue groups", len(cards))
    metric("Likely-barrier occurrences", likely_occurrences)
    metric("Review-only / informational groups", len(best_practice))
    metric("Review-only / informational occurrences", review_occurrences)
    metric("Already-triaged issue types", len(dropped))
    r += 1

    by_sev: dict[str, int] = {}
    for c in cards:
        by_sev[c.severity] = by_sev.get(c.severity, 0) + 1
    section("Likely barriers by severity")
    for sev in _SEVERITY_ORDER:
        metric(sev, by_sev.get(sev, 0))
    r += 1

    by_level: dict[str, int] = {}
    for c in cards:
        by_level[c.wcag_level or "n/a"] = by_level.get(c.wcag_level or "n/a", 0) + 1
    section("Likely barriers by WCAG level")
    for lvl in ("A", "AA", "AAA"):
        if by_level.get(lvl):
            metric(f"Level {lvl}", by_level[lvl])
    r += 1

    by_principle: dict[str, int] = {}
    for c in cards:
        p = _principle_for(c.wcag_sc)
        by_principle[p] = by_principle.get(p, 0) + 1
    section("Likely barriers by WCAG principle (POUR)")
    for principle in ("Perceivable", "Operable", "Understandable", "Robust"):
        if by_principle.get(principle):
            metric(principle, by_principle[principle])
    r += 1

    cov = coverage_matrix.summary()
    section("WCAG 2.2 A/AA coverage (all 55 criteria)")
    for m in coverage_matrix.METHODS:
        metric(coverage_matrix.METHOD_LABELS[m], cov.by_method.get(m, 0))


# --- rollup sheets (Hotspots / Affected / Coverage) -----------------------


def _build_hotspots_sheet(ws: Worksheet, cards: list[AuditCard]) -> None:
    """Pages ranked by a severity-weighted count of the open findings on them."""
    score: dict[str, float] = {}
    count: dict[str, int] = {}
    title_for: dict[str, str | None] = {}
    for card in cards:
        w = _SEVERITY_WEIGHT.get(card.severity, 1.0)
        for loc in card.locations:
            url = loc.page_url or "(unknown)"
            score[url] = score.get(url, 0.0) + w
            count[url] = count.get(url, 0) + 1
            title_for.setdefault(url, loc.page_title)
    ranked = sorted(score.items(), key=lambda kv: -kv[1])[:MAX_HOTSPOTS]
    rows: list[tuple[Any, ...]] = [
        (url, title_for.get(url) or "", round(s), count[url]) for url, s in ranked
    ]
    _styled_table(
        ws,
        title="Page hotspots",
        subtitle=(
            "Pages carrying the most (and most severe) open findings, fix shared "
            "templates here to clear issues across the site. Weighted load = sum of "
            "severity weights (Critical 4 · Serious 3 · Moderate 2 · Minor 1)."
        ),
        headers=("Page", "Title", "Weighted load", "Findings shown"),
        rows=rows,
        widths=(60.0, 30.0, 16.0, 16.0),
        center_cols=(3, 4),
        empty_note="No open findings pinned to specific pages.",
    )
    for row_number, (url, _score) in enumerate(ranked, start=5):
        if url.startswith(("https://", "http://")):
            ws.cell(row=row_number, column=1).hyperlink = url
            ws.cell(row=row_number, column=1).style = "Hyperlink"


def _build_dom_states_sheet(
    ws: Worksheet,
    conn: sqlite3.Connection,
    scan: ExportScan,
) -> None:
    """Per-page ledger for the click-through probe.

    The operational question this answers is not "how many states" but "where
    was the sweep incomplete, and why", so ``Controls operated`` sits next to
    ``Controls found`` rather than being reported as a bare total, and the
    bound that stopped a page is spelled out in words. Pages that stopped early
    sort first: they are the ones that still need a human.
    """
    coverage = interaction_coverage.load(conn, scan.id)
    if not coverage.enabled:
        _styled_table(
            ws,
            title="DOM states, content behind a click",
            subtitle=(
                "Click-through DOM state discovery was turned off for this scan. "
                "Content that appears only after operating a menu, tab, or dialog "
                "was not tested and needs manual review."
            ),
            headers=("Page", "Controls found", "Controls operated", "States", "Why it stopped"),
            rows=[],
            widths=(58.0, 16.0, 18.0, 12.0, 44.0),
            center_cols=(2, 3, 4),
            empty_note="Probe disabled for this scan.",
        )
        return

    # Incomplete pages first, then the busiest, a reader scanning the top of
    # this sheet should be looking at the gaps, not at page one of the crawl.
    ordered = sorted(
        coverage.pages,
        key=lambda page: (not page.was_limited, -page.controls_found, page.page_url),
    )
    rows: list[tuple[Any, ...]] = [
        (
            page.page_url,
            page.controls_found,
            page.controls_operated,
            page.states,
            page.limit_text or "Swept completely",
        )
        for page in ordered
    ]
    subtitle = coverage.status_line
    if coverage.caveats:
        subtitle += "  " + "  ".join(coverage.caveats)
    _styled_table(
        ws,
        title="DOM states, content behind a click",
        subtitle=subtitle,
        headers=("Page", "Controls found", "Controls operated", "States", "Why it stopped"),
        rows=rows,
        widths=(58.0, 16.0, 18.0, 12.0, 44.0),
        center_cols=(2, 3, 4),
        empty_note=(
            "No per-page control ledger was recorded for this scan. "
            "Treat per-page click coverage as unknown rather than zero."
        ),
    )
    for row_number, page in enumerate(ordered, start=5):
        if page.page_url.startswith(("https://", "http://")):
            ws.cell(row=row_number, column=1).hyperlink = page.page_url
            ws.cell(row=row_number, column=1).style = "Hyperlink"


def _build_page_references_sheet(
    ws: Worksheet,
    scan: ExportScan,
    conn: sqlite3.Connection,
    cards: list[AuditCard],
) -> None:
    """List every issue/page pairing with plain-language element locations."""
    rows: list[tuple[Any, ...]] = []
    page_urls: list[str] = []
    asset_urls: list[str | None] = []
    cards_by_issue = {(card.pipeline, card.title): card for card in cards}
    for issue in issues.list_issues(conn, scan.id):
        detail = issues.get_issue_detail(conn, scan.id, issue.issue_key)
        if detail is None:
            continue
        card = cards_by_issue.get((issue.pipeline, issue.title))
        for page in detail.pages:
            locations = (
                [location for location in card.locations if location.page_url == page.page_url]
                if card is not None
                else [
                    location
                    for location in issue_locations(conn, issue)[0]
                    if location.page_url == page.page_url
                ]
            )
            descriptions = list(dict.fromkeys(location.description for location in locations))
            if descriptions:
                shown_descriptions = descriptions[:3]
                location_description = "\n".join(
                    f"• {description}" for description in shown_descriptions
                )
                if len(descriptions) > len(shown_descriptions):
                    location_description += (
                        f"\n…and {len(descriptions) - len(shown_descriptions)} more locations"
                    )
            else:
                location_description = "Affected page; no element-level location was recorded."
            selectors = list(
                dict.fromkeys(location.selector for location in locations if location.selector)
            )
            first_asset = next(
                (location.image_url for location in locations if location.image_url),
                None,
            )
            if selectors and any(
                selector.lstrip().startswith(("{", "[")) or len(selector) > 240
                for selector in selectors
            ):
                technical_target = (
                    "Structured engine target recorded, open the issue evidence in Axcess."
                )
            elif selectors:
                shown = selectors[:5]
                technical_target = "\n".join(shown)
                if len(selectors) > len(shown):
                    technical_target += f"\n…and {len(selectors) - len(shown)} more targets"
            else:
                technical_target = "Open image asset" if first_asset else "Not recorded"
            rows.append(
                (
                    issue.title,
                    _PIPELINE_LABEL.get(issue.pipeline, issue.pipeline),
                    page.page_url,
                    page.page_title or "",
                    location_description,
                    technical_target,
                    page.occurrence_count,
                    _dominant_status(page.status_summary),
                )
            )
            page_urls.append(page.page_url)
            asset_urls.append(first_asset)
    _styled_table(
        ws,
        title="Page references",
        subtitle=(
            "Every affected page is a direct link. Location on page describes the "
            "observed element in plain language; Technical target retains the CSS "
            "selector or a link to the image asset when available."
        ),
        headers=(
            "Issue",
            "Detection source",
            "Page",
            "Page title",
            "Location on page",
            "Technical target",
            "Occurrences",
            "Status",
        ),
        rows=rows,
        widths=(38.0, 29.0, 58.0, 30.0, 46.0, 38.0, 14.0, 24.0),
        center_cols=(7,),
        empty_note="No issue-to-page references were recorded for this scan.",
    )
    for row_number, url in enumerate(page_urls, start=5):
        if url.startswith(("https://", "http://")):
            ws.cell(row=row_number, column=3).hyperlink = url
            ws.cell(row=row_number, column=3).style = "Hyperlink"
    for row_number, asset_url in enumerate(asset_urls, start=5):
        if asset_url and asset_url.startswith(("https://", "http://")):
            ws.cell(row=row_number, column=6).hyperlink = asset_url
            ws.cell(row=row_number, column=6).style = "Hyperlink"


def _build_affected_sheet(ws: Worksheet, cards: list[AuditCard]) -> None:
    """Open issues rolled up by the user ability each one blocks."""
    by_ability: dict[str, int] = {}
    by_ability_pages: dict[str, int] = {}
    for c in cards:
        for ability in c.abilities:
            by_ability[ability] = by_ability.get(ability, 0) + 1
            by_ability_pages[ability] = by_ability_pages.get(ability, 0) + c.affected_page_count
    rows: list[tuple[Any, ...]] = [
        (_ABILITY_LABEL.get(a, a.capitalize()), n, by_ability_pages[a])
        for a, n in sorted(by_ability.items(), key=lambda kv: -kv[1])
    ]
    _styled_table(
        ws,
        title="Who is affected",
        subtitle=(
            "Each issue is tagged with the user groups it blocks. One issue can "
            "affect several groups, so these counts overlap."
        ),
        headers=("User group", "Issue types affecting them", "Across (page-instances)"),
        rows=rows,
        widths=(44.0, 30.0, 24.0),
        center_cols=(2, 3),
        empty_note="No ability metadata on the open issues.",
    )


def _build_coverage_sheet(ws: Worksheet) -> None:
    """Per-criterion coverage for all 55 WCAG 2.2 A/AA SCs (filter by Coverage)."""
    crit = sorted(
        coverage_matrix.load_matrix(),
        key=lambda c: [int(p) for p in c.sc.split(".")],
    )
    rows: list[tuple[Any, ...]] = [
        (
            c.sc,
            c.name,
            c.level,
            coverage_matrix.METHOD_LABELS[c.method],
            c.automated_check or "n/a",
            c.manual_check,
        )
        for c in crit
    ]
    _styled_table(
        ws,
        title="Coverage & method, what's automated vs. manual",
        subtitle=(
            "For every WCAG 2.2 A/AA success criterion: how Axcess covers it and "
            "what a human must still confirm. Filter Coverage = 'Manual only' to see "
            "exactly what no pipeline detects."
        ),
        headers=(
            "SC",
            "Criterion",
            "Level",
            "Coverage",
            "What Axcess does",
            "Still verify by hand",
        ),
        rows=rows,
        widths=(9.0, 34.0, 8.0, 17.0, 44.0, 44.0),
        center_cols=(1, 3),
    )


def _build_manual_review_evidence_sheet(ws: Worksheet, checks: list[dict[str, Any]]) -> None:
    """Keep manual decisions and their evidence with the XLSX handoff.

    One row is emitted for every recorded manual decision. A decision with
    several pieces of evidence gets one row per reference, which keeps page
    and external links independently reviewable and filterable.
    """
    outcome_labels = {
        "pass": "Pass",
        "fail": "Fail",
        "not_tested": "Not tested",
        "needs_follow_up": "Needs follow-up",
    }
    rows: list[tuple[Any, ...]] = []
    external_urls: list[str] = []
    for check in checks:
        outcome = str(check["outcome"])
        if outcome == "not_started":
            continue
        criterion = check["criterion"]
        evidence = check["evidence"] or [None]
        for item in evidence:
            rows.append(
                (
                    criterion["sc"],
                    criterion["name"],
                    outcome_labels.get(outcome, outcome.replace("_", " ").title()),
                    check["rationale"],
                    item.get("page_url", "") if item else "",
                    item.get("evidence_url", "") if item else "",
                    item.get("note", "") if item else "",
                )
            )
            external_urls.append(str(item.get("evidence_url") or "") if item else "")

    _styled_table(
        ws,
        title="Manual review evidence",
        subtitle=(
            "Human decisions and the page/external evidence supporting them. "
            "Rows without evidence are decisions that still need supporting references."
        ),
        headers=(
            "SC",
            "Criterion",
            "Outcome",
            "Rationale",
            "Page reference",
            "External reference",
            "Expert note",
        ),
        rows=rows,
        widths=(10.0, 34.0, 18.0, 42.0, 48.0, 48.0, 42.0),
        center_cols=(1, 3),
        empty_note="No manual decisions have been recorded yet.",
    )
    for offset, url in enumerate(external_urls, start=5):
        page_url = str(rows[offset - 5][4] or "")
        if page_url.startswith(("https://", "http://")):
            ws.cell(row=offset, column=5).hyperlink = page_url
            ws.cell(row=offset, column=5).style = "Hyperlink"
        if url.startswith(("https://", "http://")):
            ws.cell(row=offset, column=6).hyperlink = url
            ws.cell(row=offset, column=6).style = "Hyperlink"


def render_xlsx(
    scan: ExportScan,
    *,
    conn: sqlite3.Connection,
    auditor: str = "Axcess",
    audit_date: str | None = None,
    blob_store: BlobStore | None = None,
) -> bytes:
    """Render the full Excel audit workbook and return the .xlsx bytes.

    The sheet count varies with the number of issues:

    1. **Summary**, the at-a-glance dashboard (counts, severity, level,
       principle, coverage headline).
    2. **Issues Overview**, the ticket index, one row per issue, each
       linking into that issue's tab.
    3. **One tab per issue**, summary block, instance table, acceptance
       criteria (capped at ``_MAX_ISSUE_SHEETS``; the rest pool into
       **More Issues**).
    4. **Page Hotspots**, pages ranked by severity-weighted load.
    5. **Page References**, affected pages, plain-language locations, and
       retained technical targets.
    6. **Who's Affected**, open issues by the ability each blocks.
    7. **Coverage & Method**, per-criterion automated-vs-manual coverage.
    8. **Test Tracking**, the manual Pass / Fail checklist.
    9. **Manual Review Evidence**, criterion outcomes and supporting notes.

    ``conn`` is required: the issue tables are built live from the same
    ``issues.list_issues()`` / audit-report card model the Markdown report
    uses, so the two deliverables can't drift. ``audit_date`` defaults to the
    scan's finish time formatted as a human date.

    ``blob_store`` is optional: when supplied and the scan has retained
    screenshot hashes, each issue's tab embeds the first available circled
    location screenshot beside its instance table.
    """
    date_str = audit_date or _fmt_date(scan.finished_at or scan.started_at)
    cards, dropped, best_practice = build_audit_cards(conn, scan)

    wb = Workbook()
    summary = wb.active
    summary.title = "Summary"
    evaluation_record = evaluation.get_evaluation(conn, scan.id)
    _build_summary_sheet(
        summary,
        scan,
        cards,
        dropped,
        best_practice,
        audit_date=date_str,
        evaluation_record=evaluation_record,
    )

    # Names are reserved before any sheet is written: the index links to
    # them, so the link and the tab have to agree on the truncated name.
    tickets = _build_tickets(conn, scan)
    environment = _scan_environment(scan, conn)
    # Names are reserved before any sheet is written: the index links to
    # them, so the link and the tab have to agree on the truncated name.
    taken: set[str] = {"Summary", "Issues Overview", _POOLED_SHEET}
    sheet_names: list[str | None] = [
        _sheet_name(ticket, taken) if ticket.index < _MAX_ISSUE_SHEETS else None
        for ticket in tickets
    ]
    pooled = [ticket for ticket in tickets if sheet_names[ticket.index] is None]

    _build_issue_index_sheet(
        wb.create_sheet("Issues Overview"),
        scan,
        tickets,
        sheet_names,
        auditor=auditor,
        audit_date=date_str,
    )
    for ticket in tickets:
        name = sheet_names[ticket.index]
        if name is not None:
            _build_issue_detail_sheet(
                wb.create_sheet(name),
                ticket,
                environment=environment,
                blob_store=blob_store,
            )
    if pooled:
        _build_pooled_instances_sheet(wb.create_sheet(_POOLED_SHEET), pooled)
    _build_hotspots_sheet(wb.create_sheet("Page Hotspots"), cards)
    _build_page_references_sheet(wb.create_sheet("Page References"), scan, conn, cards)
    _build_dom_states_sheet(wb.create_sheet("DOM States"), conn, scan)
    _build_affected_sheet(wb.create_sheet("Who's Affected"), cards)
    _build_coverage_sheet(wb.create_sheet("Coverage & Method"))
    manual_checks = evaluation.list_manual_checks(conn, scan.id)
    manual_outcomes = {
        check["criterion"]["sc"]: {
            "pass": "Pass",
            "fail": "Fail",
            "not_tested": "Not tested",
            "needs_follow_up": "Needs follow-up",
        }.get(check["outcome"], "")
        for check in manual_checks
    }
    _build_tracking_sheet(wb.create_sheet("Test Tracking"), manual_outcomes=manual_outcomes)
    _build_manual_review_evidence_sheet(wb.create_sheet("Manual Review Evidence"), manual_checks)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
