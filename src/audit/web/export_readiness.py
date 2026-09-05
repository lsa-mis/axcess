"""Server-side readiness and labeling for public report exports.

The browser is not a security or quality boundary.  A public scan export is
therefore released as a final artifact only when its persisted expert
evaluation is complete.  Earlier evaluation states require an explicit,
per-request draft acknowledgement and receive unmistakable draft labeling.
"""

from __future__ import annotations

import csv
import io
import json
import sqlite3
from dataclasses import dataclass
from typing import Any

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill

from audit import evaluation
from audit.web import issues

_DRAFT_TITLE = "DRAFT, INCOMPLETE ACCESSIBILITY EVALUATION"
_DRAFT_EXPLANATION = (
    "This export was explicitly downloaded as a draft. Expert review is incomplete; "
    "do not treat it as a conformance determination."
)
_REVIEWED_FINDING_STATUSES = frozenset(
    {"in_progress", "remediated", "accepted_risk", "false_positive"}
)


class IncompleteEvaluationExportError(ValueError):
    """Raised when a caller has not acknowledged an incomplete evaluation."""

    def __init__(self, evaluation_status: str, blockers: list[str] | None = None) -> None:
        self.evaluation_status = evaluation_status
        self.blockers = tuple(blockers or ())
        message = (
            "The expert evaluation is not completed. To download a visibly labeled "
            "draft, set draft=acknowledged."
        )
        if self.blockers:
            message += " Final-export readiness blockers: " + " ".join(self.blockers)
        super().__init__(message)


@dataclass(frozen=True, slots=True)
class PublicExportReadiness:
    """The persisted evaluation state governing one public export response."""

    evaluation_status: str
    is_draft: bool


def assess_public_export_readiness(
    conn: sqlite3.Connection,
    scan_id: int,
    *,
    draft_acknowledged: bool,
) -> PublicExportReadiness:
    """Return the export disposition or reject an unacknowledged draft.

    The caller must establish that the scan exists and is public before this
    function runs.  Keeping that ordering avoids turning an unknown scan into
    a misleading evaluation-state conflict and preserves the separate,
    identity-protected export workflow.
    """

    record = evaluation.get_evaluation(conn, scan_id)
    evaluation_status = str(record["status"])
    readiness_blockers = evaluation.evaluation_completion_blockers(conn, scan_id, record)
    if evaluation_status == "completed":
        readiness_blockers.extend(_unreviewed_actionable_evidence_blockers(conn, scan_id))
    if evaluation_status == "completed" and not readiness_blockers:
        return PublicExportReadiness(evaluation_status=evaluation_status, is_draft=False)
    if not draft_acknowledged:
        raise IncompleteEvaluationExportError(evaluation_status, readiness_blockers)
    return PublicExportReadiness(evaluation_status=evaluation_status, is_draft=True)


def _unreviewed_actionable_evidence_blockers(
    conn: sqlite3.Connection,
    scan_id: int,
) -> list[str]:
    """Return a bounded blocker when actionable evidence lacks an expert verdict.

    ``issues.list_issues`` is the canonical confidence/lane projection.  A
    final artifact may include an automated likely-barrier group or an
    expert-review lead only after every backing finding has an explicit expert
    disposition. ``in_progress`` means the auditor confirmed an open barrier;
    the other reviewed states record remediation, accepted risk, or rejection
    as a false positive. Missing status accounting fails closed rather than
    silently promoting detector output into a final report.
    """

    unresolved_groups: list[tuple[str, int]] = []
    actionable_rows = [
        row
        for row in issues.list_issues(conn, scan_id)
        if row.review_lane in {"likely_barrier", "expert_review"}
    ]
    for row in actionable_rows:
        resolved_count = sum(
            int(count)
            for status, count in row.status_summary.items()
            if status in _REVIEWED_FINDING_STATUSES
        )
        unresolved_count = sum(
            int(count)
            for status, count in row.status_summary.items()
            if status not in _REVIEWED_FINDING_STATUSES
        )
        # The status projection should account for every backing finding. If
        # it does not, count the missing records as unresolved and fail closed.
        unresolved_count += max(len(row.finding_ids) - resolved_count - unresolved_count, 0)
        if unresolved_count:
            unresolved_groups.append((row.issue_key, unresolved_count))

    if not unresolved_groups:
        return []
    unresolved_findings = sum(count for _, count in unresolved_groups)
    preview = ", ".join(issue_key for issue_key, _ in unresolved_groups[:5])
    suffix = f", and {len(unresolved_groups) - 5} more" if len(unresolved_groups) > 5 else ""
    group_noun = "group has" if len(unresolved_groups) == 1 else "groups have"
    finding_noun = "finding" if unresolved_findings == 1 else "findings"
    return [
        f"{len(unresolved_groups)} actionable evidence {group_noun} "
        f"{unresolved_findings} unreviewed backing {finding_noun}: {preview}{suffix}. "
        "Confirm an open barrier as in progress, or set every backing finding to "
        "remediated, accepted risk, or false positive before downloading a final export."
    ]


def public_export_filename(
    scan_id: int,
    extension: str,
    readiness: PublicExportReadiness,
) -> str:
    """Return a stable final filename or an unmistakable draft filename."""

    draft_marker = "_DRAFT" if readiness.is_draft else ""
    return f"scan_{scan_id}{draft_marker}.{extension}"


def label_draft_export(
    rendered: str | bytes,
    *,
    export_format: str,
    readiness: PublicExportReadiness,
) -> str | bytes:
    """Add an in-artifact draft notice where the format supports it safely.

    Markdown, JSON, XLSX, and CSV can carry a prominent notice without making
    the artifact invalid. Draft CSVs gain one additive state column, keeping a
    rectangular, standards-compliant file that importers can explicitly map
    or ignore.
    """

    if not readiness.is_draft:
        return rendered
    status_label = readiness.evaluation_status.replace("_", " ")
    if export_format in {"markdown", "audit"}:
        if not isinstance(rendered, str):
            raise TypeError("Markdown exports must be text")
        notice = (
            f"> **{_DRAFT_TITLE}**\n>\n"
            f"> Evaluation status: **{status_label}**. {_DRAFT_EXPLANATION}\n\n"
        )
        return notice + rendered
    if export_format == "json":
        if not isinstance(rendered, str):
            raise TypeError("JSON exports must be text")
        payload: Any = json.loads(rendered)
        if not isinstance(payload, dict):
            raise ValueError("Public JSON export must contain an object")
        labeled = {
            "export_notice": {
                "label": _DRAFT_TITLE,
                "evaluation_status": readiness.evaluation_status,
                "draft": True,
                "warning": _DRAFT_EXPLANATION,
            },
            **payload,
        }
        return json.dumps(labeled, indent=2, sort_keys=True, ensure_ascii=False)
    if export_format == "xlsx":
        if not isinstance(rendered, bytes):
            raise TypeError("XLSX exports must be bytes")
        return _label_draft_workbook(rendered, status_label=status_label)
    if export_format in {"csv", "jira"}:
        if not isinstance(rendered, str):
            raise TypeError("CSV exports must be text")
        return _label_draft_csv(rendered)
    return rendered


def _label_draft_csv(rendered: str) -> str:
    """Add an explicit draft-state column while preserving valid CSV shape."""

    source = io.StringIO(rendered, newline="")
    rows = list(csv.reader(source))
    if not rows:
        return "Axcess export state\nDRAFT\n"
    output = io.StringIO(newline="")
    writer = csv.writer(output, lineterminator="\n")
    writer.writerow([*rows[0], "Axcess export state"])
    for row in rows[1:]:
        writer.writerow([*row, "DRAFT"])
    return output.getvalue()


def _label_draft_workbook(rendered: bytes, *, status_label: str) -> bytes:
    """Prepend a visible notice sheet without disturbing operational tables."""

    workbook = load_workbook(io.BytesIO(rendered))
    notice = workbook.create_sheet("DRAFT NOTICE", 0)
    notice.sheet_view.showGridLines = False
    notice.sheet_properties.tabColor = "C00000"
    notice.merge_cells("A1:F1")
    notice["A1"] = _DRAFT_TITLE
    notice["A1"].font = Font(bold=True, size=20, color="FFFFFF")
    notice["A1"].fill = PatternFill("solid", fgColor="C00000")
    notice["A1"].alignment = Alignment(vertical="center")
    notice.row_dimensions[1].height = 34
    notice["A3"] = "Evaluation status"
    notice["A3"].font = Font(bold=True)
    notice["B3"] = status_label
    notice["A5"] = "Important"
    notice["A5"].font = Font(bold=True)
    notice.merge_cells("B5:F7")
    notice["B5"] = _DRAFT_EXPLANATION
    notice["B5"].alignment = Alignment(wrap_text=True, vertical="top")
    notice.column_dimensions["A"].width = 24
    notice.column_dimensions["B"].width = 38
    for column in ("C", "D", "E", "F"):
        notice.column_dimensions[column].width = 14
    workbook.active = 0

    output = io.BytesIO()
    workbook.save(output)
    return output.getvalue()
