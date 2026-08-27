"""Integration tests for the API + SPA-shell routes.

Uses FastAPI's TestClient so no browser is needed. The legacy Jinja/HTMX
server-rendered pages were removed — the UI is now the React SPA under
``/app/`` plus this ``/api/*`` surface, blob serving, exports, favicon,
and the optional shared-token gate.
"""

from __future__ import annotations

import csv
import sqlite3
from io import BytesIO, StringIO
from pathlib import Path
from types import SimpleNamespace

import pytest
from fastapi.testclient import TestClient
from openpyxl import load_workbook

from audit import coverage_matrix, evaluation
from audit.db import repo
from audit.db.schema import connect
from audit.web import issues

pytestmark = pytest.mark.ui


# ----------------------------------------------------------- shell + health


def test_root_redirects_to_spa(client: TestClient) -> None:
    # The SPA is the only UI now; / always redirects to /app/ (which itself
    # serves a 503 "build the frontend" notice if the bundle isn't built).
    resp = client.get("/", follow_redirects=False)
    assert resp.status_code in (301, 307)
    assert resp.headers["location"] == "/app/"


def test_app_shell_redirect_when_bundle_missing(client: TestClient) -> None:
    """Without a built bundle, /app/ returns a helpful 503; with it, 200."""
    resp = client.get("/app/", follow_redirects=False)
    assert resp.status_code in (200, 503)


def test_health_returns_version(client: TestClient) -> None:
    resp = client.get("/health")
    assert resp.status_code == 200
    body = resp.json()
    assert body["status"] == "ok"
    assert "version" in body


def test_local_analysis_capability_distinguishes_bundled_ocr_and_models(
    client: TestClient, monkeypatch: pytest.MonkeyPatch
) -> None:
    from audit.web import server

    class _Response:
        def raise_for_status(self) -> None:
            return None

        def json(self) -> dict[str, object]:
            return {
                "models": [
                    {"name": "qwen3-vl:2b-instruct", "size": 1_900_000_000},
                    {"name": "gemma2:9b", "size": 5_000_000_000},
                ]
            }

    class _Client:
        async def __aenter__(self) -> _Client:
            return self

        async def __aexit__(self, *_args: object) -> None:
            return None

        async def get(self, _url: str) -> _Response:
            return _Response()

    monkeypatch.setattr(server.httpx, "AsyncClient", lambda **_kwargs: _Client())
    response = client.get("/api/capabilities/local-analysis")

    assert response.status_code == 200
    body = response.json()
    assert body["ocr"]["engine"] == "Tesseract 5"
    assert body["vision"] == {
        "available": True,
        "model": "qwen3-vl:2b-instruct",
        "installed_size_bytes": 1_900_000_000,
        "reason": None,
    }
    assert body["semantic"]["available"] is False
    assert body["semantic"]["ready_models"] == ["gemma2:9b"]
    assert body["semantic"]["missing_models"] == ["qwen2.5:7b-instruct"]


def test_favicon_svg_served(client: TestClient) -> None:
    """The brand favicon must serve from /favicon.svg with the SVG mime."""
    resp = client.get("/favicon.svg")
    assert resp.status_code == 200
    assert resp.headers["content-type"].startswith("image/svg+xml")
    # The actual mark is a UMich-blue rounded rect with a maize T — assert
    # the maize hex is present so a swap to a placeholder (or an empty
    # file) is caught loudly rather than rendering a blank tab icon.
    assert b"#FFCB05" in resp.content


def test_favicon_ico_aliased_to_svg(client: TestClient) -> None:
    """Browsers and devtools poke /favicon.ico from the root. We serve the
    same SVG payload there so every surface gets a tab icon."""
    resp = client.get("/favicon.ico")
    assert resp.status_code == 200
    assert resp.headers["content-type"].startswith("image/svg+xml")
    assert b"#FFCB05" in resp.content


# ------------------------------------------------------------------ /api/scans


def test_api_list_scans(client: TestClient, seeded_db: tuple[object, object, int]) -> None:
    _, _, scan_id = seeded_db
    resp = client.get("/api/scans")
    assert resp.status_code == 200
    body = resp.json()
    assert isinstance(body, list) and len(body) >= 1
    one = next(s for s in body if s["id"] == scan_id)
    assert one["status"] == "completed"
    assert set(one.keys()) >= {
        "id",
        "seed_url",
        "status",
        "page_count",
        "finding_count",
        "started_at",
    }


def test_api_scan_detail(client: TestClient, seeded_db: tuple[object, object, int]) -> None:
    _, _, scan_id = seeded_db
    resp = client.get(f"/api/scans/{scan_id}")
    assert resp.status_code == 200
    body = resp.json()
    assert body["id"] == scan_id
    assert "by_severity" in body
    assert set(body["by_severity"].keys()) == {
        "critical",
        "major",
        "minor",
        "info",
    }
    assert body["previous_scan_id"] is None  # only one scan seeded
    assert body["blocked"] is None
    assert body["progress"] is None  # not running
    methods = {method["key"]: method for method in body["methods_used"]}
    assert "Accessibility Conformance Testing" in methods["alfa"]["label"]
    assert "specific accessibility conditions" in methods["alfa"]["description"]
    assert "not proof" in methods["alfa"]["caveat"]
    assert methods["semantic"]["label"] == "Semantic review (local AI)"
    assert "link purpose" in methods["semantic"]["description"]
    assert methods["semantic"]["state"] == "coverage_unknown"
    assert methods["semantic"]["result"] == "Coverage not recorded for this older scan"


def test_api_scan_detail_reports_actual_method_coverage(
    client: TestClient, seeded_db: tuple[object, object, int]
) -> None:
    import json as _json
    import sqlite3 as _sqlite3

    db_path, _, _ = seeded_db
    config = _json.dumps(
        {
            "js_eager": True,
            "axe_enabled": True,
            "alfa_enabled": True,
            "ocr_enabled": True,
            "vlm_enabled": True,
            "semantic_enabled": True,
            "keyboard_probe_enabled": True,
            "responsive_checks_enabled": True,
            "method_coverage_version": 1,
        }
    )
    conn = _sqlite3.connect(str(db_path))
    try:
        cur = conn.execute(
            "INSERT INTO scans (seed_url, status, page_count, config_json, "
            "axe_pages_scanned, alfa_pages_scanned, semantic_pages_analyzed, "
            "keyboard_pages_probed, responsive_pages_probed) "
            "VALUES ('https://coverage.example/', 'completed', 2, ?, 2, 1, 2, 2, 1)",
            (config,),
        )
        scan_id = int(cur.lastrowid or 0)
        conn.execute(
            "INSERT INTO pages (scan_id, url_normalized, status_code, render_mode) "
            "VALUES (?, 'https://coverage.example/', 200, 'js')",
            (scan_id,),
        )
        conn.execute(
            "INSERT INTO pages (scan_id, url_normalized, status_code, render_mode) "
            "VALUES (?, 'https://coverage.example/about', 200, 'js')",
            (scan_id,),
        )
        conn.commit()
    finally:
        conn.close()

    response = client.get(f"/api/scans/{scan_id}")
    assert response.status_code == 200
    methods = {method["key"]: method for method in response.json()["methods_used"]}
    assert methods["axe"]["state"] == "checked"
    assert methods["axe"]["result"] == "2 pages checked"
    assert methods["alfa"]["state"] == "partial"
    assert methods["alfa"]["result"] == "1 of 2 pages checked"
    assert methods["semantic"]["state"] == "checked"
    assert methods["keyboard"]["state"] == "checked"
    assert methods["responsive"]["state"] == "partial"
    assert methods["image"]["result"] == "No images found to analyze"


def test_api_scan_detail_404(client: TestClient) -> None:
    resp = client.get("/api/scans/99999")
    assert resp.status_code == 404


def test_api_scan_detail_flags_blocked_seed(
    client: TestClient, seeded_db: tuple[object, object, int]
) -> None:
    """When the seed page returned 4xx, the detail payload surfaces a
    ``blocked`` block so the SPA can warn the user instead of silently
    showing 0 findings."""
    import sqlite3 as _sqlite3

    db_path, _, _ = seeded_db
    conn = _sqlite3.connect(str(db_path))
    conn.row_factory = _sqlite3.Row
    try:
        cur = conn.execute(
            "INSERT INTO scans (seed_url, status, page_count, finding_count, config_json) "
            "VALUES ('https://blocked.example/', 'completed', 1, 0, '{}')"
        )
        blocked_scan_id = int(cur.lastrowid or 0)
        conn.execute(
            "INSERT INTO pages (scan_id, url_normalized, status_code, title, "
            "render_mode, html_hash) VALUES (?, ?, 403, 'Just a moment...', "
            "'static', ?)",
            (blocked_scan_id, "https://blocked.example/", "0" * 64),
        )
        conn.commit()
    finally:
        conn.close()

    body = client.get(f"/api/scans/{blocked_scan_id}").json()
    assert body["blocked"] is not None
    assert body["blocked"]["status_code"] == 403


def test_api_list_findings(client: TestClient, seeded_db: tuple[object, object, int]) -> None:
    _, _, scan_id = seeded_db
    resp = client.get(f"/api/scans/{scan_id}/findings?page_size=10")
    assert resp.status_code == 200
    body = resp.json()
    assert "findings" in body
    assert body["total"] >= 1
    # Thumbnails: every listed finding carries its content_hash when blob-backed.
    hashes = {f["content_hash"] for f in body["findings"] if f["content_hash"]}
    assert len(hashes) >= 1


def test_api_finding_detail(client: TestClient, seeded_db: tuple[object, object, int]) -> None:
    resp = client.get("/api/findings/1")
    assert resp.status_code == 200
    body = resp.json()
    assert body["id"] == 1
    assert "occurrences" in body
    assert isinstance(body["occurrences"], list)


def test_api_finding_status_round_trip(
    client: TestClient, seeded_db: tuple[object, object, int]
) -> None:
    resp = client.post("/api/findings/1/status", json={"status": "reviewing"})
    assert resp.status_code == 200
    assert resp.json()["status"] == "reviewing"
    detail = client.get("/api/findings/1").json()
    assert detail["status"] == "reviewing"


def test_api_finding_status_rejects_unknown(client: TestClient) -> None:
    resp = client.post("/api/findings/1/status", json={"status": "nope"})
    assert resp.status_code == 400


def test_api_scope_preview_auto_slash(client: TestClient) -> None:
    resp = client.get("/api/scope-preview?url=https://example.com/bicentennial")
    assert resp.status_code == 200
    body = resp.json()
    assert body["path_prefix"] == "/bicentennial/"
    assert body["auto_slash_added"] is True
    assert body["normalized_url"] == "https://example.com/bicentennial/"


def test_api_scope_preview_whole_host(client: TestClient) -> None:
    body = client.get("/api/scope-preview?url=https://example.com/a&whole_host=1").json()
    assert body["whole_host"] is True
    assert body["path_prefix"] == "/"


def test_api_scope_preview_invalid(client: TestClient) -> None:
    body = client.get("/api/scope-preview?url=ftp://example.com/").json()
    assert body["error"] is not None


def test_api_create_scan_rejects_bad_url(client: TestClient) -> None:
    resp = client.post("/api/scans", json={"url": "ftp://nope/"})
    assert resp.status_code == 400
    assert "error" in resp.json()


def test_local_login_scan_rejects_non_loopback_browser(client: TestClient) -> None:
    """The certificate-free convenience flow must never become a LAN API."""

    response = client.post(
        "/api/local-login-scans",
        headers={"origin": "http://testserver"},
        json={
            "seed_url": "https://app.example.test/secure/",
            "approved_auth_origins": ["https://login.example.test"],
            "authorization_acknowledged": True,
            "max_pages": 10,
            "max_depth": 2,
            "rps": 1,
        },
    )

    assert response.status_code == 403
    assert response.json()["detail"] == (
        "Direct login scanning is restricted to this Axcess computer."
    )


def test_local_login_scan_starts_from_same_loopback_origin(
    seeded_db: tuple[Path, Path, int], monkeypatch: pytest.MonkeyPatch
) -> None:
    """The local UI can create its in-memory headed-browser handoff."""

    from audit.web import server

    db_path, blob_dir, _ = seeded_db

    captured: dict[str, object] = {}

    async def _no_browser_run(
        _db_path: object, _blob_dir: object, config: object, _run: object
    ) -> None:
        captured["config"] = config

    class _NoNetworkSession:
        def __init__(self, **_kwargs: object) -> None:
            pass

    monkeypatch.setattr(server, "_run_local_login_background", _no_browser_run)
    monkeypatch.setattr(server, "ManualAuthenticationSession", _NoNetworkSession)
    monkeypatch.setattr(
        server,
        "alfa_availability",
        lambda: SimpleNamespace(available=True, reason=None),
    )
    app = server.create_app(db_path=db_path, blob_dir=blob_dir)
    with TestClient(
        app,
        base_url="http://127.0.0.1:8765",
        client=("127.0.0.1", 45678),
    ) as local_client:
        capability = local_client.get("/api/capabilities/protected-scans")
        response = local_client.post(
            "/api/local-login-scans",
            headers={"origin": "http://127.0.0.1:8765"},
            json={
                "seed_url": "https://app.example.test/secure/",
                "approved_auth_origins": ["https://login.example.test"],
                "authorization_acknowledged": True,
                "max_pages": 10,
                "max_depth": 4,
                "rps": 0.5,
                "workers": 4,
                "whole_host": True,
                "scan_engine": "both",
                "axe_level": "AAA",
                "skip_keyboard": True,
                "skip_responsive": True,
                "skip_ocr": True,
                "skip_vlm": True,
            },
        )

    assert capability.status_code == 200
    assert capability.json()["local_available"] is True
    assert response.status_code == 201
    assert response.json()["status"] == "opening_browser"
    assert isinstance(response.json()["scan_id"], int)
    config = captured["config"]
    assert isinstance(config, server.CrawlConfig)
    assert config.max_pages == 10
    assert config.max_depth == 4
    assert config.rps == 0.5
    assert config.whole_host is True
    assert config.axe_level == "AAA"
    assert config.axe_enabled is True
    assert config.keyboard_probe_enabled is False
    assert config.responsive_checks_enabled is False
    assert config.workers == 4
    assert config.concurrency_per_host == 4
    assert config.alfa_enabled is True
    assert config.browser_only is True
    assert config.image_extraction_enabled is False


def test_local_login_scan_rejects_more_than_four_workers(client: TestClient) -> None:
    response = client.post(
        "/api/local-login-scans",
        json={
            "seed_url": "https://app.example.test/secure/",
            "authorization_acknowledged": True,
            "workers": 5,
        },
    )

    # Request validation runs before the loopback-only convenience-flow guard.
    assert response.status_code == 422


def test_local_login_scan_rejects_selected_unavailable_alfa(
    seeded_db: tuple[Path, Path, int], monkeypatch: pytest.MonkeyPatch
) -> None:
    """A selected login engine must fail loudly instead of producing zero coverage."""

    from audit.web import server

    db_path, blob_dir, _ = seeded_db
    monkeypatch.setattr(
        server,
        "alfa_availability",
        lambda: SimpleNamespace(available=False, reason="Install the pinned Alfa runner."),
    )
    app = server.create_app(db_path=db_path, blob_dir=blob_dir)
    with TestClient(
        app,
        base_url="http://127.0.0.1:8765",
        client=("127.0.0.1", 45678),
    ) as local_client:
        response = local_client.post(
            "/api/local-login-scans",
            headers={"origin": "http://127.0.0.1:8765"},
            json={
                "seed_url": "https://app.example.test/secure/",
                "authorization_acknowledged": True,
                "scan_engine": "alfa",
            },
        )

    assert response.status_code == 400
    assert response.json() == {"error": "Install the pinned Alfa runner."}


def test_local_login_scan_enables_acknowledged_local_image_analysis(
    seeded_db: tuple[Path, Path, int], monkeypatch: pytest.MonkeyPatch
) -> None:
    from audit.web import server

    db_path, blob_dir, _ = seeded_db
    captured: dict[str, object] = {}

    async def _no_browser_run(
        _db_path: object, _blob_dir: object, config: object, _run: object
    ) -> None:
        captured["config"] = config

    class _NoNetworkSession:
        def __init__(self, **_kwargs: object) -> None:
            pass

    monkeypatch.setattr(server, "_run_local_login_background", _no_browser_run)
    monkeypatch.setattr(server, "ManualAuthenticationSession", _NoNetworkSession)
    app = server.create_app(db_path=db_path, blob_dir=blob_dir)
    with TestClient(
        app,
        base_url="http://127.0.0.1:8765",
        client=("127.0.0.1", 45678),
    ) as local_client:
        response = local_client.post(
            "/api/local-login-scans",
            headers={"origin": "http://127.0.0.1:8765"},
            json={
                "seed_url": "https://app.example.test/secure/",
                "approved_auth_origins": [],
                "authorization_acknowledged": True,
                "skip_ocr": False,
                "skip_vlm": False,
                "image_analysis_acknowledged": True,
            },
        )

    assert response.status_code == 201
    config = captured["config"]
    assert isinstance(config, server.CrawlConfig)
    assert config.browser_only is True
    assert config.image_extraction_enabled is True
    assert config.ocr_enabled is True
    assert config.vlm_enabled is True
    assert config.vlm_base_url == "http://127.0.0.1:11434"


def test_local_login_image_analysis_requires_storage_acknowledgement(
    client: TestClient,
) -> None:
    response = client.post(
        "/api/local-login-scans",
        json={
            "seed_url": "https://app.example.test/secure/",
            "authorization_acknowledged": True,
            "skip_ocr": False,
            "skip_vlm": True,
            "image_analysis_acknowledged": False,
        },
    )

    # Validation happens before the loopback-only convenience-flow guard.
    assert response.status_code == 422


def test_api_create_scan_kicks_off_crawl(
    client: TestClient,
    seeded_db: tuple[object, object, int],
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    from audit.web import server as _server

    called: dict[str, object] = {}

    async def _noop(db_path, config):  # type: ignore[no-untyped-def]
        called["ran"] = True
        called["seed"] = config.seed_url

    monkeypatch.setattr(_server, "_run_background_crawl", _noop)
    resp = client.post(
        "/api/scans",
        json={
            "url": "https://example.test/",
            "max_pages": 5,
            "max_depth": 2,
            "rps": 1.0,
            "workers": 1,
            "skip_ocr": True,
            "skip_vlm": True,
        },
    )
    assert resp.status_code == 201
    body = resp.json()
    assert isinstance(body["scan_id"], int)
    import asyncio as _asyncio

    loop = _asyncio.new_event_loop()
    try:
        loop.run_until_complete(_asyncio.sleep(0.05))
    finally:
        loop.close()
    assert called.get("ran") is True


def test_background_crawl_startup_failure_finishes_scan(
    seeded_db: tuple[Path, Path, int], monkeypatch: pytest.MonkeyPatch
) -> None:
    """A pre-queue exception must not leave a permanently running report."""
    import asyncio

    from audit.web import server

    db_path, _, _ = seeded_db
    # Directory-like seeds are normalized with a trailing slash by the
    # crawler.  The up-front scan row and failure path must retain that same
    # identity or the UI can stay attached to an orphaned running row.
    seed_url = "https://startup-failure.example.test/section"
    canonical_seed = f"{seed_url}/"
    with connect(db_path) as conn:
        conn.execute(
            "INSERT INTO scans (seed_url, status, config_json) VALUES (?, 'running', '{}')",
            (canonical_seed,),
        )

    async def _fail_before_queue(_conn, _config):  # type: ignore[no-untyped-def]
        raise RuntimeError("browser could not start")

    monkeypatch.setattr(server, "run_crawl", _fail_before_queue)
    asyncio.run(
        server._run_background_crawl(
            db_path,
            server.CrawlConfig(seed_url=seed_url),
        )
    )

    with connect(db_path) as conn:
        row = conn.execute(
            "SELECT status, finished_at, failure_reason FROM scans WHERE seed_url = ?",
            (canonical_seed,),
        ).fetchone()
    assert row is not None
    assert row["status"] == "failed"
    assert row["finished_at"] is not None
    assert row["failure_reason"] == "browser could not start"


def test_prepare_scan_row_uses_crawler_seed_identity(
    seeded_db: tuple[Path, Path, int],
) -> None:
    """The progress row and crawler must not split at an auto-added slash."""

    from audit.crawler.orchestrator import _ensure_scan
    from audit.web import server

    db_path, _, _ = seeded_db
    config = server.CrawlConfig(seed_url="https://app.example.test/about")
    prepared_id = server._prepare_scan_row(db_path, config)

    with connect(db_path) as conn:
        prepared = conn.execute(
            "SELECT seed_url FROM scans WHERE id = ?", (prepared_id,)
        ).fetchone()
        crawler_id = _ensure_scan(
            conn,
            "https://app.example.test/about/",
            config,
        )
        count = conn.execute(
            "SELECT COUNT(*) AS n FROM scans WHERE seed_url LIKE 'https://app.example.test/about%'"
        ).fetchone()

    assert prepared is not None
    assert prepared["seed_url"] == "https://app.example.test/about/"
    assert crawler_id == prepared_id
    assert count is not None
    assert count["n"] == 1


def test_api_create_scan_respects_whole_host(
    client: TestClient,
    seeded_db: tuple[object, object, int],
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    """Posting whole_host + a conformance target must reach the CrawlConfig."""
    from audit.web import server as _server

    captured: dict[str, object] = {}

    async def _capture(db_path, config):  # type: ignore[no-untyped-def]
        captured["whole_host"] = config.whole_host
        captured["seed_url"] = config.seed_url
        captured["axe_level"] = config.axe_level
        captured["browser_headless"] = config.browser_headless
        captured["semantic_enabled"] = config.semantic_enabled
        captured["focus_checks_enabled"] = config.focus_checks_enabled
        captured["visual_checks_enabled"] = config.visual_checks_enabled

    monkeypatch.setattr(_server, "_run_background_crawl", _capture)
    resp = client.post(
        "/api/scans",
        json={
            "url": "https://example.test/docs",
            "max_pages": 1,
            "max_depth": 1,
            "rps": 1.0,
            "workers": 1,
            "whole_host": True,
            "show_browser": True,
            "axe_level": "AAA",
            "skip_ocr": True,
            "skip_vlm": True,
            "skip_semantic": True,
            "skip_focus": True,
            "skip_visual": True,
        },
    )
    assert resp.status_code == 201
    import asyncio as _asyncio

    loop = _asyncio.new_event_loop()
    try:
        loop.run_until_complete(_asyncio.sleep(0.05))
    finally:
        loop.close()
    assert captured.get("whole_host") is True
    assert captured.get("seed_url") == "https://example.test/docs"
    # The chosen conformance target reaches the crawl config.
    assert captured.get("axe_level") == "AAA"
    assert captured.get("browser_headless") is False
    assert captured.get("semantic_enabled") is False
    assert captured.get("focus_checks_enabled") is False
    assert captured.get("visual_checks_enabled") is False


def test_api_create_scan_selects_alfa_engine(
    client: TestClient,
    seeded_db: tuple[object, object, int],
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    """The browser-facing engine selector maps to mutually clear config flags."""
    from audit.analyzer.alfa import AlfaAvailability
    from audit.web import server as _server

    captured: dict[str, object] = {}

    async def _capture(db_path, config):  # type: ignore[no-untyped-def]
        captured["axe_enabled"] = config.axe_enabled
        captured["alfa_enabled"] = config.alfa_enabled

    monkeypatch.setattr(_server, "_run_background_crawl", _capture)
    monkeypatch.setattr(_server, "alfa_availability", lambda: AlfaAvailability(True))
    response = client.post(
        "/api/scans",
        json={
            "url": "https://example.test/docs",
            "max_pages": 1,
            "max_depth": 1,
            "rps": 1,
            "workers": 1,
            "scan_engine": "alfa",
            "skip_ocr": True,
            "skip_vlm": True,
        },
    )
    assert response.status_code == 201
    import asyncio as _asyncio

    loop = _asyncio.new_event_loop()
    try:
        loop.run_until_complete(_asyncio.sleep(0.05))
    finally:
        loop.close()
    assert captured == {"axe_enabled": False, "alfa_enabled": True}


def test_api_create_scan_rejects_unavailable_alfa(
    client: TestClient,
    seeded_db: tuple[object, object, int],
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    from audit.analyzer.alfa import AlfaAvailability
    from audit.web import server as _server

    monkeypatch.setattr(
        _server,
        "alfa_availability",
        lambda: AlfaAvailability(False, "Alfa dependencies are unavailable."),
    )
    response = client.post(
        "/api/scans", json={"url": "https://example.test/", "scan_engine": "alfa"}
    )
    assert response.status_code == 422
    assert "Alfa dependencies" in response.json()["error"]


def test_api_cancel_endpoint(client: TestClient, seeded_db: tuple[object, object, int]) -> None:
    import sqlite3 as _sqlite3

    db_path, _, _ = seeded_db
    conn = _sqlite3.connect(str(db_path))
    try:
        cur = conn.execute(
            "INSERT INTO scans (seed_url, status, page_count, finding_count, "
            "config_json) VALUES ('http://x/', 'running', 0, 0, '{}')"
        )
        running_id = int(cur.lastrowid or 0)
        conn.commit()
    finally:
        conn.close()

    resp = client.post(f"/api/scans/{running_id}/cancel")
    assert resp.status_code == 200
    assert resp.json()["ok"] is True
    # Scan should be interrupted in the DB.
    detail = client.get(f"/api/scans/{running_id}").json()
    assert detail["status"] == "interrupted"


def test_api_running_scan_includes_in_flight_and_recent(
    client: TestClient, seeded_db: tuple[object, object, int]
) -> None:
    """Running scans must serialize cleanly: pages.fetched_at is a real
    datetime and jobs.lease_until likewise. Both must round-trip through
    JSONResponse without TypeErrors. Also confirm the new in_flight_pages
    field is populated from leased jobs scoped to this scan.
    """
    import json as _json
    import sqlite3 as _sqlite3

    db_path, _, _ = seeded_db
    conn = _sqlite3.connect(str(db_path))
    try:
        cur = conn.execute(
            "INSERT INTO scans (seed_url, status, page_count, finding_count, "
            "config_json) VALUES ('http://x/', 'running', 1, 0, '{}')"
        )
        running_id = int(cur.lastrowid or 0)
        # A finished page (so recent_pages.fetched_at is a real datetime).
        conn.execute(
            "INSERT INTO pages (scan_id, url_normalized, status_code, "
            "render_mode, fetched_at) VALUES (?, ?, 200, 'static', "
            "CURRENT_TIMESTAMP)",
            (running_id, "http://x/finished"),
        )
        # A leased job (so in_flight_pages picks it up). Write lease_until in
        # the same ISO-8601 format the real queue uses — *not* SQLite's
        # CURRENT_TIMESTAMP which the default converter chokes on.
        conn.execute(
            "INSERT INTO jobs (kind, payload_json, state, lease_until) "
            "VALUES ('crawl', ?, 'leased', ?)",
            (
                _json.dumps({"url": "http://x/in-flight", "depth": 2, "scan_id": running_id}),
                "2026-04-27T16:48:53+00:00",
            ),
        )
        # And a pending one to verify the count.
        conn.execute(
            "INSERT INTO jobs (kind, payload_json, state) VALUES ('crawl', ?, 'pending')",
            (_json.dumps({"url": "http://x/queued", "depth": 1, "scan_id": running_id}),),
        )
        conn.commit()
    finally:
        conn.close()

    resp = client.get(f"/api/scans/{running_id}")
    assert resp.status_code == 200, resp.text
    body = resp.json()
    assert body["status"] == "running"
    progress = body["progress"]
    assert progress is not None
    assert progress["recent_pages"][0]["url_normalized"] == "http://x/finished"
    assert isinstance(progress["recent_pages"][0]["fetched_at"], str)
    in_flight = progress["in_flight_pages"]
    assert len(in_flight) == 1
    assert in_flight[0]["url"] == "http://x/in-flight"
    assert in_flight[0]["depth"] == 2
    assert in_flight[0]["attempts"] == 0
    assert progress["pending"] == 1
    assert progress["leased"] == 1
    assert progress["completed"] == 0
    assert progress["discovered"] == 2
    assert progress["stage"] == "scanning"
    assert progress["rendered_pages"] == 0
    assert progress["static_pages"] == 1
    assert progress["eta"] == {
        "state": "estimating",
        "min_seconds": None,
        "max_seconds": None,
        "based_on_pages": 0,
    }


def test_api_diff_404s_on_missing_compare_target(
    client: TestClient, seeded_db: tuple[object, object, int]
) -> None:
    _, _, scan_id = seeded_db
    resp = client.get(f"/api/scans/{scan_id}/diff?compare_to=99999")
    assert resp.status_code == 404


# ------------------------------------------------------------ delete cascade


def test_api_delete_scan_removes_scan_and_cascades(
    client: TestClient, seeded_db: tuple[object, object, int]
) -> None:
    """DELETE /api/scans/{id} removes the scan plus everything keyed on it.

    The fixture seeds a scan with pages, page_images, analyses, findings,
    and finding_history. After delete: the scan row is gone, the cascading
    children are gone, but the deduped images survive.
    """
    import sqlite3 as _sqlite3

    db_path, _, scan_id = seeded_db

    pre = _sqlite3.connect(str(db_path))
    pre.row_factory = _sqlite3.Row
    try:
        page_n = pre.execute(
            "SELECT COUNT(*) AS n FROM pages WHERE scan_id = ?", (scan_id,)
        ).fetchone()["n"]
        finding_n = pre.execute(
            "SELECT COUNT(*) AS n FROM findings WHERE scan_id = ?", (scan_id,)
        ).fetchone()["n"]
        image_n_before = pre.execute("SELECT COUNT(*) AS n FROM images").fetchone()["n"]
    finally:
        pre.close()
    assert page_n > 0, "fixture should seed at least one page"
    assert finding_n > 0, "fixture should seed at least one finding"

    resp = client.delete(f"/api/scans/{scan_id}")
    assert resp.status_code == 200
    body = resp.json()
    assert body == {"ok": True, "deleted_scan_id": scan_id}

    post = _sqlite3.connect(str(db_path))
    post.row_factory = _sqlite3.Row
    try:
        scan_row = post.execute("SELECT id FROM scans WHERE id = ?", (scan_id,)).fetchone()
        page_n_after = post.execute(
            "SELECT COUNT(*) AS n FROM pages WHERE scan_id = ?", (scan_id,)
        ).fetchone()["n"]
        finding_n_after = post.execute(
            "SELECT COUNT(*) AS n FROM findings WHERE scan_id = ?", (scan_id,)
        ).fetchone()["n"]
        history_n_after = post.execute(
            "SELECT COUNT(*) AS n FROM finding_history WHERE scan_id = ?",
            (scan_id,),
        ).fetchone()["n"]
        image_n_after = post.execute("SELECT COUNT(*) AS n FROM images").fetchone()["n"]
        dangling = post.execute(
            "SELECT COUNT(*) AS n FROM images WHERE first_seen_scan_id = ?",
            (scan_id,),
        ).fetchone()["n"]
    finally:
        post.close()

    assert scan_row is None
    assert page_n_after == 0
    assert finding_n_after == 0
    assert history_n_after == 0
    assert image_n_after == image_n_before
    assert dangling == 0


def test_api_delete_scan_404s_for_unknown_id(client: TestClient) -> None:
    resp = client.delete("/api/scans/9999999")
    assert resp.status_code == 404


def test_api_delete_scan_clears_jobs_for_that_scan(
    client: TestClient, seeded_db: tuple[object, object, int]
) -> None:
    """Jobs aren't FK-bound (scan_id lives in payload_json), so the delete
    handler must explicitly remove them."""
    import sqlite3 as _sqlite3

    db_path, _, scan_id = seeded_db
    conn = _sqlite3.connect(str(db_path))
    conn.row_factory = _sqlite3.Row
    try:
        conn.execute(
            "INSERT INTO jobs (kind, payload_json, state) VALUES ('fetch', ?, 'pending')",
            (f'{{"url":"http://x/a","scan_id":{scan_id},"depth":0}}',),
        )
        conn.execute(
            "INSERT INTO jobs (kind, payload_json, state) VALUES ('fetch', ?, 'failed')",
            (f'{{"url":"http://x/b","scan_id":{scan_id},"depth":0}}',),
        )
        conn.execute(
            "INSERT INTO scans (seed_url, status, page_count, finding_count, "
            "config_json) VALUES ('http://other.example/', 'completed', 0, 0, '{}')"
        )
        other_scan = int(conn.execute("SELECT last_insert_rowid() AS i").fetchone()["i"])
        conn.execute(
            "INSERT INTO jobs (kind, payload_json, state) VALUES ('fetch', ?, 'pending')",
            (f'{{"url":"http://x/c","scan_id":{other_scan},"depth":0}}',),
        )
        conn.commit()
    finally:
        conn.close()

    resp = client.delete(f"/api/scans/{scan_id}")
    assert resp.status_code == 200

    check = _sqlite3.connect(str(db_path))
    check.row_factory = _sqlite3.Row
    try:
        for_deleted = check.execute(
            "SELECT COUNT(*) AS n FROM jobs WHERE json_extract(payload_json, '$.scan_id') = ?",
            (scan_id,),
        ).fetchone()["n"]
        for_other = check.execute(
            "SELECT COUNT(*) AS n FROM jobs WHERE json_extract(payload_json, '$.scan_id') = ?",
            (other_scan,),
        ).fetchone()["n"]
    finally:
        check.close()

    assert for_deleted == 0, "jobs for the deleted scan should be removed"
    assert for_other == 1, "jobs for sibling scans must not be touched"


def test_api_delete_scan_409s_when_scan_is_running(
    client: TestClient, seeded_db: tuple[object, object, int]
) -> None:
    """A scan with an active asyncio task must not be deleted out from
    under the worker. The handler returns 409 and the scan stays put."""
    import asyncio
    import sqlite3 as _sqlite3

    db_path, _, _ = seeded_db
    conn = _sqlite3.connect(str(db_path))
    conn.row_factory = _sqlite3.Row
    try:
        cur = conn.execute(
            "INSERT INTO scans (seed_url, status, page_count, finding_count, "
            "config_json) VALUES ('http://live.example/', 'running', 0, 0, '{}')"
        )
        running_id = int(cur.lastrowid or 0)
        conn.commit()
    finally:
        conn.close()

    # Find the handler's captured crawl_state dict by inspecting the
    # endpoint's closure cells (matched by structure so a rename is safe).
    app = client.app  # type: ignore[attr-defined]
    target: dict[str, object] | None = None
    for route in app.routes:  # type: ignore[attr-defined]
        endpoint = getattr(route, "endpoint", None)
        if endpoint is None or endpoint.__name__ != "api_delete_scan":
            continue
        for cell in endpoint.__closure__ or ():
            val = cell.cell_contents
            if isinstance(val, dict) and "task" in val and "scan_id" in val:
                target = val
                break
        if target is not None:
            break
    assert target is not None, "could not locate crawl_state via closure"

    async def _never() -> None:
        await asyncio.sleep(60)

    loop = asyncio.new_event_loop()
    try:
        task = loop.create_task(_never())
        target["task"] = task
        target["scan_id"] = running_id
        try:
            resp = client.delete(f"/api/scans/{running_id}")
        finally:
            task.cancel()
            loop.run_until_complete(asyncio.gather(task, return_exceptions=True))
            target["task"] = None
            target["scan_id"] = None
    finally:
        loop.close()

    assert resp.status_code == 409
    detail = resp.json().get("detail", "")
    assert "running" in detail.lower()

    check = _sqlite3.connect(str(db_path))
    try:
        row = check.execute("SELECT status FROM scans WHERE id = ?", (running_id,)).fetchone()
    finally:
        check.close()
    assert row is not None
    assert row[0] == "running"


def test_stale_running_scans_interrupted_on_server_boot(tmp_path) -> None:  # type: ignore[no-untyped-def]
    """When the app boots it marks every DB row stuck in 'running' as
    'interrupted', because the live task driving it is gone after a
    process restart."""
    import sqlite3 as _sqlite3
    from pathlib import Path as _Path

    from audit.db.schema import connect as _connect

    db_path = tmp_path / "sweep.db"
    migrations_dir = _Path(__file__).resolve().parents[2] / "src" / "audit" / "db" / "migrations"
    prep = _connect(db_path)
    try:
        for p in sorted(migrations_dir.glob("*.sql")):
            if p.name.endswith(".rollback.sql"):
                continue
            prep.executescript(p.read_text())
        prep.execute(
            "INSERT INTO scans (seed_url, status, config_json) "
            "VALUES ('http://stuck.example/', 'running', '{}')"
        )
        prep.execute(
            "INSERT INTO scans (seed_url, status, config_json) "
            "VALUES ('http://done.example/', 'completed', '{}')"
        )
    finally:
        prep.close()

    from audit.web.server import create_app

    create_app(db_path=db_path, blob_dir=tmp_path / "blobs")

    check = _sqlite3.connect(str(db_path))
    check.row_factory = _sqlite3.Row
    try:
        rows = check.execute("SELECT seed_url, status FROM scans ORDER BY id").fetchall()
    finally:
        check.close()
    states = {r["seed_url"]: r["status"] for r in rows}
    assert states["http://stuck.example/"] == "interrupted"
    assert states["http://done.example/"] == "completed"


# --------------------------------------------------------------------- /api/tracking


def test_api_tracking_returns_shipped_and_roadmap(client: TestClient) -> None:
    """The coverage tracker endpoint reports what's shipped vs. planned,
    sourced from coverage_status.py."""
    resp = client.get("/api/tracking")
    assert resp.status_code == 200
    body = resp.json()
    # Shipped pipelines include the axe rule engine and the VLM image one.
    pipelines = {p["pipeline"] for p in body["shipped"]}
    assert {"axe", "image", "semantic", "keyboard", "responsive"} <= pipelines
    # Roadmap carries WCAG-keyed rows with the three-way status.
    by_wcag = {r["wcag"]: r for r in body["roadmap"]}
    assert by_wcag["1.4.5"]["status"] == "shipped"
    assert by_wcag["2.4.4"]["status"] == "shipped"
    # 1.3.2 shipped via the visual (VLM) probe this session.
    assert by_wcag["1.3.2"]["status"] == "shipped"
    # 3.2.3 still needs the cross-page embedding analyzer — not started.
    assert by_wcag["3.2.3"]["status"] == "planned"
    assert body["counts"]["shipped"] >= 2

    # The WCAG 2.2 A/AA coverage matrix rides along (the transparency model).
    cov = body["coverage"]
    assert cov["total"] == 55
    assert cov["covered"] + cov["manual_only"] == cov["total"]
    assert sum(cov["by_method"].values()) == cov["total"]
    # Every criterion carries the "what to test manually" promise.
    assert all(c["manual_check"] for c in cov["criteria"])
    # A known shipped SC is non-manual and names its pipeline.
    kb = next(c for c in cov["criteria"] if c["sc"] == "2.1.2")
    assert kb["method"] != "manual" and "keyboard" in kb["pipelines"]


# --------------------------------------------------------------- blob serving


def test_blob_serves_png_bytes(client: TestClient, seeded_db: tuple[object, object, int]) -> None:
    # Pull a blob-backed finding's content_hash from the API, then fetch it.
    _, _, scan_id = seeded_db
    findings = client.get(f"/api/scans/{scan_id}/findings?page_size=10").json()["findings"]
    content_hash = next(f["content_hash"] for f in findings if f["content_hash"])
    resp = client.get(f"/blobs/{content_hash}")
    assert resp.status_code == 200
    assert resp.headers["content-type"].startswith("image/")
    assert resp.content[:8] == b"\x89PNG\r\n\x1a\n"


def test_blob_rejects_invalid_hash(client: TestClient) -> None:
    resp = client.get("/blobs/not_a_hash")
    assert resp.status_code == 400


def test_blob_returns_404_for_unknown_hash(client: TestClient) -> None:
    resp = client.get("/blobs/" + "f" * 64)
    assert resp.status_code == 404


# ------------------------------------------------------------------- exports
# Exports are served only under /api/* now (the React SPA downloads via a
# plain <a download> to that URL). The legacy /scans/{id}/export alias is
# gone with the rest of the Jinja routes.


def _set_actionable_finding_statuses(
    conn: sqlite3.Connection,
    scan_id: int,
    status: str,
) -> None:
    for row in issues.list_issues(conn, scan_id):
        if row.review_lane == "informational":
            continue
        rationale = f"Expert reviewed {row.issue_key} and recorded {status}."
        if row.pipeline == "image":
            repo.bulk_set_findings_status(
                conn,
                finding_ids=row.finding_ids,
                status=status,
                rationale=rationale,
            )
        else:
            repo.bulk_set_a11y_findings_status(
                conn,
                finding_ids=row.finding_ids,
                status=status,
                rationale=rationale,
            )


def _complete_evaluation(
    client: TestClient,
    db_path: object,
    scan_id: int,
    *,
    actionable_status: str | None = "remediated",
) -> None:
    conn = connect(Path(db_path))
    try:
        record = evaluation.upsert_evaluation(
            conn,
            scan_id,
            {
                "reviewer": "A. Expert",
                "purpose": "Prepare an accessibility remediation handoff.",
                "scope_included": "Pages captured in this scan.",
                "methods_note": "Automated evidence plus criterion-by-criterion expert review.",
                "limitations": "Results are limited to the recorded pages and review date.",
                "status": "in_progress",
            },
        )
        conn.executemany(
            "INSERT INTO manual_check_results "
            "(evaluation_report_id, criterion_sc, outcome, rationale, tested_at) "
            "VALUES (?, ?, 'pass', ?, CURRENT_TIMESTAMP)",
            (
                (
                    int(record["id"]),
                    criterion.sc,
                    f"Expert reviewed {criterion.sc} against the documented scope.",
                )
                for criterion in coverage_matrix.load_matrix()
            ),
        )
        if actionable_status is not None:
            _set_actionable_finding_statuses(conn, scan_id, actionable_status)
    finally:
        conn.close()
    response = client.put(
        f"/api/scans/{scan_id}/evaluation",
        json={"status": "completed"},
    )
    assert response.status_code == 200, response.text


def test_export_csv_route(client: TestClient, seeded_db: tuple[object, object, int]) -> None:
    db_path, _, scan_id = seeded_db
    _complete_evaluation(client, db_path, scan_id)
    resp = client.get(f"/api/scans/{scan_id}/export/csv")
    assert resp.status_code == 200
    assert "text/csv" in resp.headers["content-type"]
    assert "attachment" in resp.headers["content-disposition"]
    assert f"scan_{scan_id}.csv" in resp.headers["content-disposition"]
    assert "DRAFT" not in resp.headers["content-disposition"]
    assert "x-axcess-export-state" not in resp.headers
    assert "finding_id,severity" in resp.text


def test_export_json_route(client: TestClient, seeded_db: tuple[object, object, int]) -> None:
    db_path, _, scan_id = seeded_db
    _complete_evaluation(client, db_path, scan_id)
    resp = client.get(f"/api/scans/{scan_id}/export/json")
    assert resp.status_code == 200
    assert "application/json" in resp.headers["content-type"]
    body = resp.json()
    assert "export_notice" not in body
    assert body["scan"]["id"] == scan_id
    assert isinstance(body["findings"], list)


def test_export_jira_route(client: TestClient, seeded_db: tuple[object, object, int]) -> None:
    db_path, _, scan_id = seeded_db
    _complete_evaluation(client, db_path, scan_id)
    resp = client.get(f"/api/scans/{scan_id}/export/jira")
    assert resp.status_code == 200
    assert "text/csv" in resp.headers["content-type"]
    assert "Summary,Description,Priority" in resp.text


def test_export_markdown_route(client: TestClient, seeded_db: tuple[object, object, int]) -> None:
    db_path, _, scan_id = seeded_db
    _complete_evaluation(client, db_path, scan_id)
    resp = client.get(f"/api/scans/{scan_id}/export/markdown")
    assert resp.status_code == 200
    assert "text/markdown" in resp.headers["content-type"]
    assert resp.text.startswith(f"# Accessibility evidence inventory — Scan #{scan_id}")


def test_export_xlsx_route(client: TestClient, seeded_db: tuple[object, object, int]) -> None:
    db_path, _, scan_id = seeded_db
    _complete_evaluation(client, db_path, scan_id)
    resp = client.get(f"/api/scans/{scan_id}/export/xlsx")
    assert resp.status_code == 200
    assert "spreadsheetml.sheet" in resp.headers["content-type"]
    assert f'filename="scan_{scan_id}.xlsx"' in resp.headers["content-disposition"]
    assert resp.content[:2] == b"PK"  # a real .xlsx (zip) body
    assert load_workbook(BytesIO(resp.content)).sheetnames[0] == "Summary"


@pytest.mark.parametrize("fmt", ("csv", "json", "jira", "markdown", "audit", "xlsx"))
def test_incomplete_evaluation_export_requires_explicit_draft_acknowledgement(
    client: TestClient,
    seeded_db: tuple[object, object, int],
    fmt: str,
) -> None:
    _, _, scan_id = seeded_db

    response = client.get(f"/api/scans/{scan_id}/export/{fmt}")

    assert response.status_code == 409
    assert "expert evaluation is not completed" in response.json()["detail"]
    assert "draft=acknowledged" in response.json()["detail"]


@pytest.mark.parametrize("fmt", ("csv", "json", "jira", "markdown", "audit", "xlsx"))
def test_acknowledged_draft_export_uses_draft_filename_and_visible_label(
    client: TestClient,
    seeded_db: tuple[object, object, int],
    fmt: str,
) -> None:
    _, _, scan_id = seeded_db

    response = client.get(
        f"/api/scans/{scan_id}/export/{fmt}",
        params={"draft": "acknowledged"},
    )

    assert response.status_code == 200, response.text
    assert f"scan_{scan_id}_DRAFT." in response.headers["content-disposition"]
    assert response.headers["x-axcess-export-state"] == "draft"
    if fmt in {"markdown", "audit"}:
        assert response.text.startswith("> **DRAFT — INCOMPLETE ACCESSIBILITY EVALUATION**")
        assert "do not treat it as a conformance determination" in response.text
    elif fmt == "json":
        notice = response.json()["export_notice"]
        assert notice["draft"] is True
        assert notice["evaluation_status"] == "draft"
        assert notice["label"] == "DRAFT — INCOMPLETE ACCESSIBILITY EVALUATION"
    elif fmt == "xlsx":
        workbook = load_workbook(BytesIO(response.content))
        assert workbook.sheetnames[0] == "DRAFT NOTICE"
        assert workbook["DRAFT NOTICE"]["A1"].value == (
            "DRAFT — INCOMPLETE ACCESSIBILITY EVALUATION"
        )
        assert workbook["DRAFT NOTICE"]["B3"].value == "draft"
    elif fmt == "csv":
        assert response.text.startswith("finding_kind,finding_id,")
        assert response.text.splitlines()[0].endswith(",Axcess export state")
        assert response.text.splitlines()[1].endswith(",DRAFT")
    else:
        assert response.text.startswith("Summary,Description,Priority,")
        rows = list(csv.reader(StringIO(response.text)))
        assert rows[0][-1] == "Axcess export state"
        assert all(row[-1] == "DRAFT" for row in rows[1:])


@pytest.mark.parametrize("draft_value", (None, "true", "Acknowledged", "ACKNOWLEDGED"))
def test_only_exact_draft_acknowledgement_parameter_unlocks_an_incomplete_export(
    client: TestClient,
    seeded_db: tuple[object, object, int],
    draft_value: str | None,
) -> None:
    _, _, scan_id = seeded_db
    params = {} if draft_value is None else {"draft": draft_value}

    response = client.get(
        f"/api/scans/{scan_id}/export/csv",
        params=params,
        headers={"X-Axcess-Acknowledge-Draft": "true"},
    )

    assert response.status_code == 409


def test_unreviewed_actionable_findings_block_final_but_allow_acknowledged_draft(
    client: TestClient,
    seeded_db: tuple[object, object, int],
) -> None:
    db_path, _, scan_id = seeded_db
    conn = connect(Path(db_path))
    try:
        page_id = int(
            conn.execute(
                "SELECT id FROM pages WHERE scan_id = ? ORDER BY id LIMIT 1", (scan_id,)
            ).fetchone()["id"]
        )
        repo.upsert_axe_violation(
            conn,
            page_id=page_id,
            scan_id=scan_id,
            rule_id="label",
            wcag_sc="1.3.1",
            wcag_scs="1.3.1,3.3.2",
            wcag_level="A",
            impact="serious",
            help="Form elements must have labels",
            help_url="https://dequeuniversity.com/rules/axe/4.10/label",
            target_selector="#email",
            failure_summary="Fix the missing label.",
            html_snippet='<input id="email">',
            target_hash="export-readiness-likely-barrier",
        )
    finally:
        conn.close()
    _complete_evaluation(
        client,
        db_path,
        scan_id,
        actionable_status=None,
    )

    final = client.get(f"/api/scans/{scan_id}/export/markdown")
    assert final.status_code == 409
    assert "actionable evidence" in final.json()["detail"]
    assert "unreviewed backing" in final.json()["detail"]
    assert "axe:label" in final.json()["detail"]

    draft = client.get(
        f"/api/scans/{scan_id}/export/markdown",
        params={"draft": "acknowledged"},
    )
    assert draft.status_code == 200
    assert f"scan_{scan_id}_DRAFT.md" in draft.headers["content-disposition"]
    assert draft.text.startswith("> **DRAFT — INCOMPLETE ACCESSIBILITY EVALUATION**")


@pytest.mark.parametrize(
    "reviewed_status", ("in_progress", "remediated", "accepted_risk", "false_positive")
)
def test_reviewed_actionable_findings_allow_final_export(
    client: TestClient,
    seeded_db: tuple[object, object, int],
    reviewed_status: str,
) -> None:
    db_path, _, scan_id = seeded_db
    _complete_evaluation(
        client,
        db_path,
        scan_id,
        actionable_status=reviewed_status,
    )

    response = client.get(f"/api/scans/{scan_id}/export/markdown")

    assert response.status_code == 200
    assert f'filename="scan_{scan_id}.md"' in response.headers["content-disposition"]
    assert "DRAFT" not in response.headers["content-disposition"]


def test_export_unknown_format_rejected(
    client: TestClient, seeded_db: tuple[object, object, int]
) -> None:
    _, _, scan_id = seeded_db
    resp = client.get(f"/api/scans/{scan_id}/export/xml")
    assert resp.status_code == 400


def test_export_unknown_scan_returns_404(client: TestClient) -> None:
    resp = client.get("/api/scans/99999/export/csv")
    assert resp.status_code == 404


def test_legacy_jinja_export_alias_is_gone(client: TestClient) -> None:
    """The old /scans/{id}/export/{fmt} Jinja alias was removed; only the
    /api/* route remains. The bare path now falls through to the SPA
    catch-all-less router and 404s."""
    _ = client
    resp = client.get("/scans/1/export/csv")
    assert resp.status_code == 404


# --------------------------------------------------------------------
# Hosting: the optional shared-token gate (docs/hosting.md, Path A).
# Routes are checked against /api/scans (a real gated route) now that the
# Jinja /scans page is gone; /health stays open.
# --------------------------------------------------------------------


def test_access_token_unset_is_no_op(
    seeded_db: tuple[object, object, int],
    monkeypatch,  # type: ignore[no-untyped-def]
) -> None:
    """With no AUDIT_ACCESS_TOKEN, every route is open (the local default)."""
    from pathlib import Path

    from audit.web.server import create_app

    monkeypatch.delenv("AUDIT_ACCESS_TOKEN", raising=False)
    db_path, blob_dir, _ = seeded_db
    app = create_app(db_path=Path(db_path), blob_dir=Path(blob_dir))
    c = TestClient(app)
    assert c.get("/api/scans").status_code == 200
    assert c.get("/health").status_code == 200


def test_access_token_gate_blocks_and_admits(
    seeded_db: tuple[object, object, int],
    monkeypatch,  # type: ignore[no-untyped-def]
) -> None:
    """When the token is set: 401 without it, 200 with it, /health always open."""
    from pathlib import Path

    from audit.web.server import create_app

    monkeypatch.setenv("AUDIT_ACCESS_TOKEN", "s3cret-token")
    db_path, blob_dir, _ = seeded_db
    app = create_app(db_path=Path(db_path), blob_dir=Path(blob_dir))
    c = TestClient(app)

    # No token → 401.
    assert c.get("/api/scans").status_code == 401

    # Health stays open (uptime checks don't carry the token).
    assert c.get("/health").status_code == 200

    # Query-string token → 200, and a cookie is set for next time.
    ok = c.get("/api/scans?token=s3cret-token")
    assert ok.status_code == 200
    assert "aa_access" in ok.cookies

    # Header forms also work (API/CLI clients).
    fresh = TestClient(app)
    assert fresh.get("/api/scans", headers={"X-Access-Token": "s3cret-token"}).status_code == 200
    fresh2 = TestClient(app)
    assert (
        fresh2.get("/api/scans", headers={"Authorization": "Bearer s3cret-token"}).status_code
        == 200
    )

    # Wrong token → 401.
    bad = TestClient(app)
    assert bad.get("/api/scans?token=nope").status_code == 401
