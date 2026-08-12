"""Expert-review persistence, exports, and read-only report-tool boundaries."""

from __future__ import annotations

import sqlite3
from io import BytesIO

from openpyxl import load_workbook

from audit import evaluation
from audit.db import repo
from audit.exports.audit_report import render_audit_report
from audit.exports.collector import collect_scan
from audit.exports.xlsx_export import render_xlsx
from audit.mcp_server import get_finding_evidence, get_report_summary, list_report_issues


def _scan(conn: sqlite3.Connection, *, url: str = "https://example.test/") -> int:
    cur = conn.execute(
        "INSERT INTO scans (seed_url, status, config_json) VALUES (?, 'completed', '{}')",
        (url,),
    )
    return int(cur.lastrowid)


def test_evaluation_and_manual_check_are_persisted_without_mutating_scan(
    tmp_db: sqlite3.Connection,
) -> None:
    scan_id = _scan(tmp_db)
    initial = evaluation.get_evaluation(tmp_db, scan_id)
    assert initial["exists"] is False
    assert initial["target_standard"] == "WCAG 2.2"

    record = evaluation.upsert_evaluation(
        tmp_db,
        scan_id,
        {
            "reviewer": "A. Expert",
            "scope_included": "Homepage and admissions",
            "status": "in_progress",
        },
    )
    assert record["exists"] is True
    assert record["reviewer"] == "A. Expert"
    assert (
        tmp_db.execute("SELECT status FROM scans WHERE id = ?", (scan_id,)).fetchone()["status"]
        == "completed"
    )

    result = evaluation.update_manual_check(
        tmp_db,
        scan_id=scan_id,
        criterion_sc="2.1.1",
        outcome="fail",
        rationale="Keyboard-only menu cannot reach the search control.",
    )
    assert result["outcome"] == "fail"
    assert result["rationale"].startswith("Keyboard-only")


def test_manual_matrix_exposes_the_protected_authentication_review_prompt(
    tmp_db: sqlite3.Connection,
) -> None:
    """SC 3.3.8 is a deliberate expert check, not a browser-session result."""
    scan_id = _scan(tmp_db)

    authentication = next(
        check
        for check in evaluation.list_manual_checks(tmp_db, scan_id)
        if check["criterion"]["sc"] == "3.3.8"
    )

    assert authentication["outcome"] == "not_started"
    assert authentication["criterion"]["method"] == "manual"
    assert "temporary browser session" in authentication["criterion"]["manual_check"]


def test_outcome_only_manual_projection_never_returns_review_prose(
    tmp_db: sqlite3.Connection,
) -> None:
    """A protected boundary can retain a result without exposing evidence."""

    scan_id = _scan(tmp_db)
    evaluation.update_manual_check(
        tmp_db,
        scan_id=scan_id,
        criterion_sc="3.3.8",
        outcome="fail",
        rationale="Sensitive sign-in observation must not leave this boundary.",
    )
    evaluation.add_manual_evidence(
        tmp_db,
        scan_id=scan_id,
        criterion_sc="3.3.8",
        note="Sensitive authentication detail.",
        page_id=None,
        evidence_url="https://sensitive.example.test/",
    )

    protected_result = evaluation.update_manual_check_outcome_only(
        tmp_db,
        scan_id=scan_id,
        criterion_sc="3.3.8",
        outcome="needs_follow_up",
    )
    listed = next(
        check
        for check in evaluation.list_manual_check_outcomes(tmp_db, scan_id)
        if check["criterion"]["sc"] == "3.3.8"
    )

    assert protected_result == listed
    assert set(listed) == {"criterion", "outcome", "tested_at", "updated_at"}
    assert listed["outcome"] == "needs_follow_up"
    assert "rationale" not in listed
    assert "evidence" not in listed
    row = tmp_db.execute(
        """
        SELECT r.rationale
          FROM manual_check_results r
          JOIN evaluation_reports e ON e.id = r.evaluation_report_id
         WHERE e.scan_id = ? AND r.criterion_sc = '3.3.8'
        """,
        (scan_id,),
    ).fetchone()
    assert row is not None and row["rationale"] == ""

    # The protected-safe projection must not even read the public evaluation
    # prose or manual-evidence table. This catches a future convenience call
    # to ``get_evaluation`` / ``list_manual_checks`` at the boundary.
    statements: list[str] = []
    tmp_db.set_trace_callback(statements.append)
    try:
        evaluation.list_manual_check_outcomes(tmp_db, scan_id)
    finally:
        tmp_db.set_trace_callback(None)
    observed_sql = "\n".join(statements).lower()
    assert "manual_check_evidence" not in observed_sql
    assert "purpose" not in observed_sql
    assert "scope_included" not in observed_sql
    assert "rationale" not in observed_sql


def test_manual_evidence_rejects_page_from_another_report(tmp_db: sqlite3.Connection) -> None:
    scan_a = _scan(tmp_db, url="https://a.test/")
    scan_b = _scan(tmp_db, url="https://b.test/")
    page_b = tmp_db.execute(
        "INSERT INTO pages (scan_id, url_normalized, render_mode) "
        "VALUES (?, ?, 'static') RETURNING id",
        (scan_b, "https://b.test/"),
    ).fetchone()["id"]

    try:
        evaluation.add_manual_evidence(
            tmp_db,
            scan_id=scan_a,
            criterion_sc="1.1.1",
            note="Wrong report page",
            page_id=int(page_b),
            evidence_url="",
        )
    except ValueError as exc:
        assert "not part" in str(exc)
    else:  # pragma: no cover - assertion branch
        raise AssertionError("cross-report evidence must be rejected")


def test_expert_record_is_rendered_in_report_and_tracking_sheet(tmp_db: sqlite3.Connection) -> None:
    scan_id = _scan(tmp_db)
    evaluation.upsert_evaluation(
        tmp_db,
        scan_id,
        {"reviewer": "A. Expert", "limitations": "No authenticated flows were reviewed."},
    )
    evaluation.update_manual_check(
        tmp_db,
        scan_id=scan_id,
        criterion_sc="1.1.1",
        outcome="pass",
        rationale="Representative meaningful images have text alternatives.",
    )
    scan = collect_scan(tmp_db, scan_id)
    markdown = render_audit_report(scan, conn=tmp_db)
    assert "## Expert evaluation record" in markdown
    assert "A. Expert" in markdown
    assert "No authenticated flows" in markdown
    assert "| 1.1.1 | pass |" in markdown

    book = load_workbook(BytesIO(render_xlsx(scan, conn=tmp_db)))
    tracking = book["Test Tracking"]
    first_row = next(
        row
        for row in tracking.iter_rows(min_row=5, values_only=True)
        if str(row[0]).startswith("1.1.1")
    )
    assert first_row[2] == "Pass"


def test_xlsx_tracking_preserves_non_pass_manual_outcomes(
    tmp_db: sqlite3.Connection,
) -> None:
    """Handoff must not flatten honest uncertainty into an empty workbook cell."""
    scan_id = _scan(tmp_db)
    evaluation.update_manual_check(
        tmp_db,
        scan_id=scan_id,
        criterion_sc="2.1.1",
        outcome="not_tested",
        rationale="Keyboard session has not happened.",
    )
    evaluation.update_manual_check(
        tmp_db,
        scan_id=scan_id,
        criterion_sc="1.4.10",
        outcome="needs_follow_up",
        rationale="Reflow needs a second viewport review.",
    )

    scan = collect_scan(tmp_db, scan_id)
    book = load_workbook(BytesIO(render_xlsx(scan, conn=tmp_db)))
    tracking = book["Test Tracking"]
    outcomes = {
        str(row[0]).split(" ", 1)[0]: row[2]
        for row in tracking.iter_rows(min_row=5, values_only=True)
    }

    assert outcomes["2.1.1"] == "Not tested"
    assert outcomes["1.4.10"] == "Needs follow-up"


def test_xlsx_handoff_keeps_scope_and_manual_evidence_linked(
    tmp_db: sqlite3.Connection,
) -> None:
    """Excel contains the same human context and evidence as the audit report."""
    scan_id = _scan(tmp_db)
    page_id = repo.upsert_page(
        tmp_db,
        scan_id=scan_id,
        url_normalized="https://example.test/",
        status_code=200,
        title="Home",
        render_mode="js",
        html_hash="0" * 64,
    )
    evaluation.upsert_evaluation(
        tmp_db,
        scan_id,
        {
            "scope_included": "Public pages",
            "limitations": "Authenticated paths excluded.",
            "reviewer": "A. Expert",
        },
    )
    evaluation.update_manual_check(
        tmp_db,
        scan_id=scan_id,
        criterion_sc="2.1.1",
        outcome="fail",
        rationale="The menu cannot be completed by keyboard.",
    )
    evaluation.add_manual_evidence(
        tmp_db,
        scan_id=scan_id,
        criterion_sc="2.1.1",
        page_id=page_id,
        evidence_url="https://tracker.example/A11Y-42",
        note="Keyboard session recording.",
    )

    book = load_workbook(BytesIO(render_xlsx(collect_scan(tmp_db, scan_id), conn=tmp_db)))
    summary = book["Summary"]
    summary_values = {
        str(summary.cell(row=row, column=1).value or ""): summary.cell(row=row, column=2).value
        for row in range(1, summary.max_row + 1)
    }
    assert summary_values["Included scope"] == "Public pages"
    assert summary_values["Limitations"] == "Authenticated paths excluded."

    evidence = book["Manual Review Evidence"]
    assert tuple(evidence.cell(row=4, column=column).value for column in range(1, 8)) == (
        "SC",
        "Criterion",
        "Outcome",
        "Rationale",
        "Page reference",
        "External reference",
        "Expert note",
    )
    row = next(row for row in evidence.iter_rows(min_row=5) if row[0].value == "2.1.1")
    assert row[2].value == "Fail"
    assert row[3].value == "The menu cannot be completed by keyboard."
    assert row[4].value == "https://example.test/"
    assert row[5].hyperlink is not None
    assert row[5].hyperlink.target == "https://tracker.example/A11Y-42"
    assert row[6].value == "Keyboard session recording."


def test_read_only_report_tools_reject_cross_report_finding(tmp_db: sqlite3.Connection) -> None:
    scan_a = _scan(tmp_db, url="https://a.test/")
    scan_b = _scan(tmp_db, url="https://b.test/")
    assert get_report_summary(tmp_db, scan_id=scan_a)["report"]["id"] == scan_a
    assert list_report_issues(tmp_db, scan_id=scan_a) == []
    assert get_finding_evidence(tmp_db, scan_id=scan_a, finding_id=9999) is None
    assert get_finding_evidence(tmp_db, scan_id=scan_b, finding_id=9999) is None
