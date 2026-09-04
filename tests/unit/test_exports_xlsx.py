"""Excel (.xlsx) audit report — structural assertions.

The xlsx renderer produces a binary workbook, so we parse it back with
openpyxl and pin the *structure* (two named sheets, the metadata header
block, the table headers, value mappings) rather than byte content. The
issue rows come from the same multi-pipeline fixture the Markdown audit
report uses, so both deliverables stay in lockstep.
"""

from __future__ import annotations

import io
import re
import sqlite3
from typing import Any

import pytest
from openpyxl import load_workbook
from test_audit_report import _scan_with_real_findings  # tests/unit on sys.path

from audit import coverage_matrix
from audit.exports import xlsx_export
from audit.exports.collector import collect_scan
from audit.exports.xlsx_export import (
    _INSTANCE_HEADERS,
    _ISSUE_HEADERS,
    _TRACK_HEADERS,
    render_xlsx,
)


def _render(conn: sqlite3.Connection) -> bytes:
    scan_id = _scan_with_real_findings(conn)
    scan = collect_scan(conn, scan_id, ui_base_url="http://127.0.0.1:8765")
    return render_xlsx(scan, conn=conn)


def test_xlsx_is_a_valid_workbook_with_all_report_sheets(tmp_db: sqlite3.Connection) -> None:
    data = _render(tmp_db)
    assert data[:2] == b"PK"  # zip / OOXML signature
    wb = load_workbook(io.BytesIO(data))
    names = wb.sheetnames
    assert names[0] == "Summary"
    assert names[1] == "Issues Overview"
    # Per-issue tabs sit between the index and the rollups, so the rollup
    # sheets are located by name rather than by a fixed position.
    assert names[-6:] == [
        "Page Hotspots",
        "Page References",
        "Who's Affected",
        "Coverage & Method",
        "Test Tracking",
        "Manual Review Evidence",
    ]
    issue_tabs = names[2:-6]
    assert issue_tabs, "every issue should get its own tab"
    assert all(re.fullmatch(r"I\d{2}( .*)?", name) for name in issue_tabs)
    # Excel rejects longer names outright, so this is a hard limit.
    assert all(len(name) <= 31 for name in names)


def test_summary_dashboard_has_metadata_and_rollups(tmp_db: sqlite3.Connection) -> None:
    wb = load_workbook(io.BytesIO(_render(tmp_db)))
    ws = wb["Summary"]
    # Flatten the label/value column-A/B grid into a lookup.
    cells = {
        str(ws.cell(row=r, column=1).value or ""): ws.cell(row=r, column=2).value
        for r in range(1, ws.max_row + 1)
    }
    assert cells.get("Audited against") == "WCAG 2.2 Level AA"
    assert cells.get("Pages crawled") == 2
    # Section bars + at least one severity + coverage-method row are present.
    labels = set(cells)
    assert "Likely barriers by severity" in labels  # a section header (col A, no value)
    assert "Likely-barrier issue groups" in labels
    assert "Review-only / informational groups" in labels
    assert "Manual only" in labels  # coverage-method rollup row
    assert {"Critical", "Serious", "Moderate", "Minor"} <= labels


def test_coverage_sheet(tmp_db: sqlite3.Connection) -> None:
    wb = load_workbook(io.BytesIO(_render(tmp_db)))

    cov = wb["Coverage & Method"]
    assert cov.cell(row=4, column=1).value == "SC"
    cov_rows = [r for r in range(5, cov.max_row + 1) if cov.cell(row=r, column=1).value]
    assert len(cov_rows) == len(coverage_matrix.load_matrix())
    methods = {str(cov.cell(row=r, column=4).value or "") for r in cov_rows}
    assert "Manual only" in methods


def test_issues_overview_indexes_every_issue_and_links_to_its_tab(
    tmp_db: sqlite3.Connection,
) -> None:
    wb = load_workbook(io.BytesIO(_render(tmp_db)))
    ws = wb["Issues Overview"]

    # Metadata block (rows 1-5), merged into column A.
    meta = [str(ws.cell(row=r, column=1).value or "") for r in range(1, 6)]
    assert meta[0].startswith("Page: http://example.com/")
    assert meta[1] == "Compliance Standard: WCAG 2.2 Level AA"
    assert meta[2].startswith("Audit Date:")
    assert meta[3] == "Auditor: Axcess"
    assert "Prioritization Guidance" in meta[4]

    header = tuple(ws.cell(row=8, column=c).value for c in range(1, len(_ISSUE_HEADERS) + 1))
    assert header == _ISSUE_HEADERS

    data_rows = [r for r in range(9, ws.max_row + 1) if ws.cell(row=r, column=1).value]
    assert data_rows
    ids = [str(ws.cell(row=r, column=1).value) for r in data_rows]
    assert ids == [f"I{n:02d}" for n in range(1, len(ids) + 1)]

    # Instance and page counts are numbers, not prose, so the index sorts.
    assert all(isinstance(ws.cell(row=r, column=7).value, int) for r in data_rows)
    assert all(isinstance(ws.cell(row=r, column=8).value, int) for r in data_rows)

    # Every row links to a tab that actually exists in this workbook.
    for row in data_rows:
        link = ws.cell(row=row, column=len(_ISSUE_HEADERS)).hyperlink
        assert link is not None
        assert link.target is None, "issue links stay inside the workbook"
        sheet = str(link.location).split("!")[0].strip("'")
        assert sheet in wb.sheetnames

    conformance = {ws.cell(row=r, column=4).value for r in data_rows}
    assert "S" in conformance  # page-has-heading-one is best-practice → S
    assert conformance & {"A", "AA"}  # real SCs present too

    references = wb["Page References"]
    reference_headers = tuple(references.cell(row=4, column=c).value for c in range(1, 9))
    assert reference_headers == (
        "Issue",
        "Detection source",
        "Page",
        "Page title",
        "Location on page",
        "Technical target",
        "Occurrences",
        "Status",
    )
    descriptions = {
        str(references.cell(row=r, column=5).value or "") for r in range(5, references.max_row + 1)
    }
    assert any(
        "text element containing “low text” with class “muted”" in value for value in descriptions
    )
    technical_targets = {
        str(references.cell(row=r, column=6).value or "") for r in range(5, references.max_row + 1)
    }
    assert "p > span.muted" in technical_targets
    assert references.cell(row=5, column=3).hyperlink is not None


def test_issue_tab_carries_the_summary_block_and_one_row_per_instance(
    tmp_db: sqlite3.Connection,
) -> None:
    """The ticket itself: what the issue is, then every place it was seen."""
    wb = load_workbook(io.BytesIO(_render(tmp_db)))
    index = wb["Issues Overview"]
    first_tab = str(index.cell(row=9, column=len(_ISSUE_HEADERS)).hyperlink.location)
    ws = wb[first_tab.split("!")[0].strip("'")]

    assert str(ws.cell(row=1, column=1).value or "").startswith("I01 · ")
    back = ws.cell(row=2, column=1)
    assert back.hyperlink is not None
    assert str(back.hyperlink.location).startswith("'Issues Overview'")

    labels = {
        str(ws.cell(row=r, column=1).value or ""): ws.cell(row=r, column=2).value
        for r in range(4, ws.max_row + 1)
    }
    assert str(labels.get("Current Behavior") or "").strip()
    assert str(labels.get("Severity") or "").strip()
    assert str(labels.get("Impact") or "").strip()
    # A number, not prose: the index column has to sort by it.
    assert isinstance(labels.get("Instances"), int)
    assert isinstance(labels.get("Pages"), int)

    header_row = next(r for r in range(4, ws.max_row + 1) if ws.cell(row=r, column=1).value == "#")
    header = tuple(ws.cell(row=header_row, column=c).value for c in range(1, 7))
    assert header == _INSTANCE_HEADERS

    numbers = []
    for r in range(header_row + 1, ws.max_row + 1):
        value = ws.cell(row=r, column=1).value
        if not isinstance(value, int):
            break
        numbers.append(value)
        assert str(ws.cell(row=r, column=2).value or "").strip(), "instance has no location"
        assert str(ws.cell(row=r, column=3).value or "").strip(), "instance has no user action"
    assert numbers == list(range(1, len(numbers) + 1))
    # The count in the summary block is the real occurrence count, which can
    # exceed the rows listed; it must never be smaller than them.
    assert int(labels["Instances"]) >= len(numbers)


def _issue_sheet_labels(sheet: Any) -> dict[str, str]:
    """Column A label -> column B value for one issue tab's summary block."""
    labels: dict[str, str] = {}
    for row in sheet.iter_rows():
        key = str(row[0].value or "").strip()
        if key and len(row) > 1 and row[1].value is not None:
            labels.setdefault(key, str(row[1].value))
    return labels


def test_category_and_environment_are_derived_never_left_blank(
    tmp_db: sqlite3.Connection,
) -> None:
    """Both fields carry real content, and Environment admits what it lacks.

    Category is derived from the criterion, so it always resolves to one of
    the WCAG principle buckets the Summary sheet already counts. Environment
    states only conditions the crawl can prove, and ends with an explicit
    slot for the framework and assistive tech a human has to supply — a
    blank there would read as "nothing more to say".
    """
    wb = load_workbook(io.BytesIO(_render(tmp_db)))
    buckets = {"Perceivable", "Operable", "Understandable", "Robust", "Other / best-practice"}
    sheets = [s for s in wb.worksheets if s.title.startswith("I0")]
    assert sheets, "expected at least one issue tab"
    for sheet in sheets:
        labels = _issue_sheet_labels(sheet)
        assert labels["Category"] in buckets
        environment = labels["Environment"]
        assert "page(s) crawled" in environment
        assert "add before sharing" in environment, (
            "Environment must mark what only a human can fill in"
        )


def test_suggested_fix_appears_only_where_options_were_authored(
    tmp_db: sqlite3.Connection,
) -> None:
    """No empty Option/Approach table, and no options invented from fix steps.

    An empty table would read as "there is no way to fix this"; a single row
    synthesized from the fix steps would read as "these were the
    alternatives considered". The section is present only when someone wrote
    the approaches into rules/audit_report.yaml.
    """
    wb = load_workbook(io.BytesIO(_render(tmp_db)))
    for sheet in wb.worksheets:
        rows = list(sheet.iter_rows())
        headings = [str(row[0].value or "").strip() for row in rows]
        if "Suggested fix" not in headings:
            continue
        start = headings.index("Suggested fix")
        header = next(
            (row for row in rows[start:] if str(row[0].value or "").strip() == "Option"),
            None,
        )
        assert header is not None, f"{sheet.title}: Suggested fix without a table header"
        assert [str(cell.value or "") for cell in header[:4]] == [
            "Option",
            "Applies to",
            "Approach",
            "Watch out for",
        ]
        body = rows[rows.index(header) + 1]
        assert str(body[0].value or "").strip(), f"{sheet.title}: Suggested fix table is empty"
        assert str(body[2].value or "").strip(), f"{sheet.title}: an option with no approach"


def test_xlsx_without_blob_store_embeds_no_evidence_images(
    tmp_db: sqlite3.Connection,
) -> None:
    """Without a blob store there is nothing to embed, and nothing is faked."""
    scan_id = _scan_with_real_findings(tmp_db)
    scan = collect_scan(tmp_db, scan_id, ui_base_url="http://127.0.0.1:8765")
    data = render_xlsx(scan, conn=tmp_db, blob_store=None)
    assert data[:2] == b"PK"
    wb = load_workbook(io.BytesIO(data))
    assert all(not sheet._images for sheet in wb.worksheets)
    assert wb["Issues Overview"].cell(row=8, column=1).value == "ID"


def test_page_references_summarizes_large_structured_engine_targets(
    tmp_db: sqlite3.Connection,
) -> None:
    """Alfa JSON evidence must not turn one workbook row thousands of pixels tall."""
    scan_id = _scan_with_real_findings(tmp_db)
    page_id = int(
        tmp_db.execute(
            "SELECT id FROM pages WHERE scan_id = ? ORDER BY id LIMIT 1",
            (scan_id,),
        ).fetchone()["id"]
    )
    structured_target = '{"path":[' + ",".join(f'"node-{index}"' for index in range(150)) + "]}"
    tmp_db.execute(
        """
        INSERT INTO page_a11y_findings
            (page_id, scan_id, pipeline, engine_outcome, rule_id, wcag_sc,
             wcag_scs, wcag_level, impact, help, help_url, target_selector,
             failure_summary, html_snippet, target_hash, status)
        VALUES (?, ?, 'alfa', 'failed', 'sia-r-test', '1.1.1', '1.1.1', 'A',
                NULL, 'Alfa structured target test',
                'https://alfa.siteimprove.com/rules/sia-r-test', ?,
                'The ACT rule failed for the stored target.', NULL,
                'structured-target-hash', 'new')
        """,
        (page_id, scan_id, structured_target),
    )
    scan = collect_scan(tmp_db, scan_id, ui_base_url="http://127.0.0.1:8765")
    workbook = load_workbook(io.BytesIO(render_xlsx(scan, conn=tmp_db)))
    references = workbook["Page References"]
    alfa_rows = [
        row
        for row in range(5, references.max_row + 1)
        if "Siteimprove Alfa" in str(references.cell(row=row, column=2).value or "")
    ]
    assert alfa_rows
    row = alfa_rows[0]
    assert "structured target evidence" in str(references.cell(row=row, column=5).value)
    assert references.cell(row=row, column=6).value == (
        "Structured engine target recorded — open the issue evidence in Axcess."
    )


def test_test_tracking_is_the_full_manual_checklist(tmp_db: sqlite3.Connection) -> None:
    wb = load_workbook(io.BytesIO(_render(tmp_db)))
    ws = wb["Test Tracking"]

    header = tuple(ws.cell(row=4, column=c).value for c in range(1, len(_TRACK_HEADERS) + 1))
    assert header == _TRACK_HEADERS

    focus = [str(ws.cell(row=r, column=1).value or "") for r in range(5, ws.max_row + 1)]
    # One row per WCAG 2.2 A/AA criterion, sorted, each with a check.
    assert len(focus) == len(coverage_matrix.load_matrix())
    assert any(f.startswith("1.1.1 ") for f in focus)
    assert all("What to check" not in f for f in focus)
    # Every checklist row carries a non-empty "what to check" instruction.
    checks = [ws.cell(row=r, column=2).value for r in range(5, ws.max_row + 1)]
    assert all(str(c or "").strip() for c in checks)


def test_every_web_url_in_workbook_is_a_clickable_excel_link(
    tmp_db: sqlite3.Connection,
) -> None:
    """No handoff recipient should need to copy and paste a visible web URL."""
    workbook = load_workbook(io.BytesIO(_render(tmp_db)))
    url_pattern = re.compile(r"https?://[^\s]+")

    for sheet in workbook.worksheets:
        for row in sheet.iter_rows():
            for cell in row:
                value = str(cell.value or "")
                if url_pattern.search(value):
                    assert cell.hyperlink is not None, (
                        f"{sheet.title}!{cell.coordinate} contains a plain-text URL: {value}"
                    )


def test_issues_past_the_tab_cap_are_pooled_not_dropped(
    tmp_db: sqlite3.Connection, monkeypatch: pytest.MonkeyPatch
) -> None:
    """A big scan must stay openable without losing an issue.

    The cap only moves where instances live. Every issue keeps its index row,
    so the workbook never quietly reports fewer issues than the scan found.
    """
    monkeypatch.setattr(xlsx_export, "_MAX_ISSUE_SHEETS", 1)
    scan_id = _scan_with_real_findings(tmp_db)
    scan = collect_scan(tmp_db, scan_id, ui_base_url="http://127.0.0.1:8765")
    wb = load_workbook(io.BytesIO(render_xlsx(scan, conn=tmp_db)))

    assert "More Issues" in wb.sheetnames
    issue_tabs = [name for name in wb.sheetnames if re.fullmatch(r"I\d{2}( .*)?", name)]
    assert len(issue_tabs) == 1

    index = wb["Issues Overview"]
    rows = [r for r in range(9, index.max_row + 1) if index.cell(row=r, column=1).value]
    assert len(rows) > 1, "the index must still list every issue"
    pooled_links = [
        index.cell(row=r, column=len(_ISSUE_HEADERS)).hyperlink.location for r in rows[1:]
    ]
    assert all(str(link).startswith("'More Issues'") for link in pooled_links)

    pooled = wb["More Issues"]
    pooled_ids = {
        str(pooled.cell(row=r, column=1).value or "") for r in range(5, pooled.max_row + 1)
    }
    indexed_ids = {str(index.cell(row=r, column=1).value) for r in rows[1:]}
    # Every pooled issue that has retained locations appears in the sheet.
    assert pooled_ids & indexed_ids
